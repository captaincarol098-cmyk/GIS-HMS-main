from pathlib import Path
from uuid import UUID
from datetime import datetime
from fastapi import APIRouter, BackgroundTasks, Depends, File, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook, load_workbook
from io import BytesIO
from shapely.geometry import Point, shape
from ..database import get_db
from ..models import BulkImportJob, User, Child, Purok, Barangay
from ..models.entities import Sex, JobStatus
from ..middleware.rbac import get_current_user

router = APIRouter(prefix="/api/import", tags=["import"])
UPLOAD_DIR = Path("uploads")


def point_in_purok(lat: float, lng: float, purok_geometry: dict) -> bool:
    """Check if a point (lat, lng) is inside a purok's polygon."""
    try:
        point = Point(lng, lat)  # shapely uses (lng, lat) order
        polygon = shape(purok_geometry)
        return polygon.contains(point)
    except:
        return False


async def process_excel_file(job_id: UUID, file_path: str, user_id: UUID, db: AsyncSession):
    """Background task to process the uploaded Excel file."""
    try:
        # Load job
        job = await db.get(BulkImportJob, job_id)
        if not job:
            return
        
        job.status = JobStatus.processing
        await db.commit()
        
        # Load Excel file
        wb = load_workbook(file_path)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        job.total_rows = len(rows)
        await db.commit()
        
        # Load all puroks with geometry for auto-detection
        puroks_stmt = select(Purok).where(Purok.is_archived.is_(False))
        puroks = (await db.scalars(puroks_stmt)).all()
        
        # Load barangays for name lookup
        barangays_stmt = select(Barangay)
        barangays = {b.name: b for b in (await db.scalars(barangays_stmt)).all()}
        
        errors = []
        success_count = 0
        
        for idx, row in enumerate(rows, start=2):
            try:
                (full_name, birth_date, sex, guardian_name, contact_number, 
                 barangay_name, purok_name, latitude, longitude, 
                 measurement_date, weight_kg, height_cm, muac_cm) = row
                
                # Validate required fields
                if not full_name or not birth_date or not sex:
                    errors.append(f"Row {idx}: Missing required fields (name, birth_date, sex)")
                    continue
                
                # Find barangay
                barangay = barangays.get(barangay_name)
                if not barangay:
                    errors.append(f"Row {idx}: Barangay '{barangay_name}' not found")
                    continue
                
                # Auto-detect purok from coordinates if provided
                purok_id = None
                if latitude and longitude:
                    try:
                        lat = float(latitude)
                        lng = float(longitude)
                        
                        # Find purok that contains this point
                        for purok in puroks:
                            if purok.barangay_id == barangay.id and purok.geometry:
                                if point_in_purok(lat, lng, purok.geometry):
                                    purok_id = purok.id
                                    break
                        
                        if not purok_id:
                            # If no purok found by geometry, try name match
                            if purok_name:
                                purok_by_name = next((p for p in puroks if p.name == purok_name and p.barangay_id == barangay.id), None)
                                if purok_by_name:
                                    purok_id = purok_by_name.id
                    except ValueError:
                        errors.append(f"Row {idx}: Invalid coordinates")
                        lat = lng = None
                else:
                    lat = lng = None
                
                # Create child record
                child = Child(
                    full_name=full_name,
                    birth_date=birth_date if isinstance(birth_date, datetime) else datetime.strptime(str(birth_date), "%Y-%m-%d").date(),
                    sex=Sex(sex.lower()),
                    guardian_name=guardian_name,
                    contact_number=contact_number,
                    barangay_id=barangay.id,
                    purok_id=purok_id,
                    latitude=lat,
                    longitude=lng,
                    is_active=True
                )
                db.add(child)
                await db.flush()
                
                success_count += 1
                
            except Exception as e:
                errors.append(f"Row {idx}: {str(e)}")
        
        # Update job status
        job.status = JobStatus.completed
        job.success_rows = success_count
        job.error_rows = len(errors)
        job.errors = errors
        job.completed_at = datetime.now()
        await db.commit()
        
    except Exception as e:
        # Update job with error
        if job:
            job.status = JobStatus.failed
            job.errors = [f"Processing failed: {str(e)}"]
            await db.commit()


@router.post("/upload")
async def upload(background_tasks: BackgroundTasks, file: UploadFile = File(...), db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    UPLOAD_DIR.mkdir(exist_ok=True)
    path = UPLOAD_DIR / file.filename
    path.write_bytes(await file.read())
    job = BulkImportJob(uploaded_by=user.id, file_name=file.filename, file_path=str(path))
    db.add(job)
    await db.commit()
    await db.refresh(job)
    
    # Start background processing
    background_tasks.add_task(process_excel_file, job.id, str(path), user.id, db)
    
    return {"job_id": str(job.id), "status": job.status.value}


@router.get("/template")
async def template():
    wb = Workbook()
    ws = wb.active
    ws.append(["full_name", "birth_date", "sex", "guardian_name", "contact_number", "barangay_name", "purok_name", "latitude", "longitude", "measurement_date", "weight_kg", "height_cm", "muac_cm"])
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=child_import_template.xlsx"})


@router.get("/jobs")
async def jobs(db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    return list((await db.scalars(select(BulkImportJob).order_by(BulkImportJob.created_at.desc()))).all())


@router.get("/jobs/{job_id}")
async def job(job_id: UUID, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    return await db.get(BulkImportJob, job_id)
