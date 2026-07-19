"""
OPT Plus API Router
Endpoints for OPT Plus calculations and 3-tier alert system
"""

from fastapi import APIRouter, Depends, HTTPException, Query, File, UploadFile
from pydantic import BaseModel, Field
from typing import Optional, Literal
from uuid import UUID
import uuid
from datetime import date, datetime
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from openpyxl import load_workbook
from io import BytesIO
from ..database import get_db
from ..middleware.rbac import get_current_user
from ..models import User, Child, Measurement, Barangay, Purok
from ..models.entities import Sex, WazStatus, HazStatus, WhzStatus
from ..services.opt_plus_calculations import (
    calculate_age_in_months,
    adjust_height_for_position,
    calculate_full_opt_plus_assessment
)
from ..services.three_tier_alerts import (
    generate_comprehensive_alert,
    generate_all_barangay_alerts,
    generate_all_purok_alerts,
    calculate_prevalence_rates,
    determine_alert_level,
    classify_risk_level
)

router = APIRouter(prefix="/api/opt-plus", tags=["OPT Plus"])

# ============================================================================
# REQUEST/RESPONSE MODELS
# ============================================================================

class OPTPlusCalculationRequest(BaseModel):
    sex: Literal["male", "female"]
    birth_date: date
    measurement_date: date
    weight_kg: float = Field(gt=0, description="Weight in kilograms")
    height_cm: float = Field(gt=0, description="Height/length in centimeters")
    measurement_position: Literal["standing", "lying_down"] = "standing"
    has_bilateral_edema: bool = False


class AgeCalculationRequest(BaseModel):
    birth_date: date
    measurement_date: date


class HeightAdjustmentRequest(BaseModel):
    height_cm: float
    age_months: int
    measurement_position: Literal["standing", "lying_down"]


class ImportMeasurementRow(BaseModel):
    child_id: Optional[UUID] = None
    full_name: str
    birth_date: date
    sex: Literal["M", "F", "male", "female"]
    measurement_date: date
    weight_kg: float
    height_cm: float
    muac_cm: Optional[float] = None
    notes: Optional[str] = None


# ============================================================================
# MODULE 1: AGE CALCULATION
# ============================================================================

@router.post("/calculate-age")
async def calculate_age(
    body: AgeCalculationRequest,
    user: User = Depends(get_current_user)
):
    """
    Calculate age in months using NNC OPT Plus Guidelines [50]
    
    ⚠️ NOTE: OPT Plus features are disabled for admin and superadmin roles.
    
    Algorithm: Chronological Milestone Resolution
    Formula: Age (in months) = FLOOR((Measurement Date - Birth Date) / 30.4375)
    
    Rules:
    - Disregard remaining days (floor function)
    """
    age_months = calculate_age_in_months(body.birth_date, body.measurement_date)
    
    return {
        "birth_date": body.birth_date,
        "measurement_date": body.measurement_date,
        "age_months": age_months,
        "age_years": age_months // 12,
        "remaining_months": age_months % 12
    }


# ============================================================================
# MODULE 2: HEIGHT/LENGTH ADJUSTMENT
# ============================================================================

@router.post("/adjust-height")
async def adjust_height(
    body: HeightAdjustmentRequest,
    user: User = Depends(get_current_user)
):
    """
    Adjust height/length based on measurement position per NNC OPT Plus Guidelines [50]
    
    ⚠️ NOTE: OPT Plus features are disabled for admin and superadmin roles.
    
    Rules:
    - 0-23 months + Standing (wrong): Add +0.7 cm
    - 24-71 months + Lying (wrong): Subtract -0.7 cm
    - Round to nearest 0.5 cm
    """
    adjusted_height = adjust_height_for_position(
        body.height_cm,
        body.age_months,
        body.measurement_position
    )
    
    adjustment = adjusted_height - round(body.height_cm * 2) / 2
    
    return {
        "raw_height_cm": body.height_cm,
        "rounded_height_cm": round(body.height_cm * 2) / 2,
        "adjusted_height_cm": adjusted_height,
        "adjustment_cm": round(adjustment, 1),
        "age_months": body.age_months,
        "measurement_position": body.measurement_position,
        "age_group": "0-23 months" if body.age_months < 24 else "24-71 months",
        "correct_position": "lying_down" if body.age_months < 24 else "standing"
    }


# ============================================================================
# MODULE 3-5: COMPLETE OPT PLUS ASSESSMENT
# ============================================================================

@router.post("/calculate-assessment")
async def calculate_assessment(
    body: OPTPlusCalculationRequest,
    user: User = Depends(get_current_user)
):
    """
    Complete OPT Plus assessment with all modules
    
    ⚠️ NOTE: OPT Plus features are disabled for admin and superadmin roles.
    
    Includes:
    1. Age calculation
    2. Height adjustment
    3. Z-score calculation (LMS method)
    4. Edema override check
    5. Nutritional status classification
    """
    try:
        result = calculate_full_opt_plus_assessment(
            sex=body.sex,
            birth_date=body.birth_date,
            measurement_date=body.measurement_date,
            weight_kg=body.weight_kg,
            height_cm=body.height_cm,
            measurement_position=body.measurement_position,
            has_bilateral_edema=body.has_bilateral_edema
        )
        
        return {
            "success": True,
            "input": body.dict(),
            "assessment": result
        }
    except ValueError as e:
        raise HTTPException(400, str(e))


# ============================================================================
# MODULE 6: 3-TIER ALERT GENERATION
# ============================================================================

@router.get("/alerts/comprehensive")
async def get_comprehensive_alert(
    barangay_id: Optional[UUID] = Query(None),
    purok_id: Optional[UUID] = Query(None),
    include_trends: bool = Query(False),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """
    Generate comprehensive 3-tier alert with all analytics
    
    ⚠️ NOTE: OPT Plus features are disabled for admin and superadmin roles.
    
    Alert Levels:
    - Warning (🟡 Yellow): Prevalence ≥ 10%
    - Critical (🟠 Orange): Wasting ≥ 5% OR Any Indicator ≥ 15%
    - Emergency (🔴 Red): Wasting ≥ 10% OR SAM ≥ 2%
    """
    # BHW can only access their barangay
    if user.role.value == "admin":
        barangay_id = user.barangay_id
    
    alert = await generate_comprehensive_alert(
        db, barangay_id, purok_id, include_trends
    )
    
    return alert


@router.get("/alerts/barangays")
async def get_all_barangay_alerts(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """
    Generate alerts for all barangays
    ⚠️ NOTE: OPT Plus features are disabled for admin and superadmin roles.
    """
    alerts = await generate_all_barangay_alerts(db)
    
    # Filter for BHW users
    if user.role.value == "admin":
        alerts = [a for a in alerts if a.get("barangay_id") == str(user.barangay_id)]
    
    return {
        "alerts": alerts,
        "total_barangays": len(alerts),
        "emergency_count": sum(1 for a in alerts if a["alert"]["alert_level"] == "emergency"),
        "critical_count": sum(1 for a in alerts if a["alert"]["alert_level"] == "critical"),
        "warning_count": sum(1 for a in alerts if a["alert"]["alert_level"] == "warning")
    }


@router.get("/alerts/puroks")
async def get_all_purok_alerts(
    barangay_id: Optional[UUID] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """
    Generate alerts for all puroks
    ⚠️ NOTE: OPT Plus features are disabled for admin and superadmin roles.
    """
    # BHW can only access their barangay
    if user.role.value == "admin":
        barangay_id = user.barangay_id
    
    alerts = await generate_all_purok_alerts(db, barangay_id)
    
    return {
        "alerts": alerts,
        "barangay_id": str(barangay_id) if barangay_id else None,
        "total_puroks": len(alerts),
        "emergency_count": sum(1 for a in alerts if a["alert"]["alert_level"] == "emergency"),
        "critical_count": sum(1 for a in alerts if a["alert"]["alert_level"] == "critical"),
        "warning_count": sum(1 for a in alerts if a["alert"]["alert_level"] == "warning")
    }


# ============================================================================
# MODULE 8-9: PREVALENCE AND TREND ANALYSIS
# ============================================================================

@router.get("/prevalence")
async def get_prevalence_rates(
    barangay_id: Optional[UUID] = Query(None),
    purok_id: Optional[UUID] = Query(None),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """
    Calculate malnutrition prevalence rates
    ⚠️ NOTE: OPT Plus features are disabled for admin and superadmin roles.
    
    Formula:
    Prevalence (%) = (Number of children with condition / Total children assessed) × 100
    
    Target Age Group: Children aged 0-71 months
    """
    # BHW can only access their barangay
    if user.role.value == "admin":
        barangay_id = user.barangay_id
    
    prevalence = await calculate_prevalence_rates(
        db, barangay_id, purok_id, start_date, end_date
    )
    
    alert = determine_alert_level(prevalence)
    risk = classify_risk_level(prevalence)
    
    return {
        "prevalence": prevalence,
        "alert": alert,
        "risk": risk,
        "filters": {
            "barangay_id": str(barangay_id) if barangay_id else None,
            "purok_id": str(purok_id) if purok_id else None,
            "start_date": start_date,
            "end_date": end_date
        }
    }


# ============================================================================
# MODULE 10: HOTSPOT DETECTION
# ============================================================================

@router.get("/hotspots")
async def get_hotspots(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """
    Detect malnutrition hotspots
    ⚠️ NOTE: OPT Plus features are disabled for admin and superadmin roles.
    
    Risk Levels:
    - High Risk (Red): ≥ 30% prevalence
    - Medium Risk (Yellow): 15% - 29% prevalence
    - Low Risk (Green): < 15% prevalence
    """
    alerts = await generate_all_barangay_alerts(db)
    
    # Filter for BHW users
    if user.role.value == "admin":
        alerts = [a for a in alerts if a.get("barangay_id") == str(user.barangay_id)]
    
    # Classify hotspots
    hotspots = {
        "high_risk": [],
        "medium_risk": [],
        "low_risk": []
    }
    
    for alert in alerts:
        risk_level = alert["risk"]["risk_level"]
        if risk_level == "high":
            hotspots["high_risk"].append(alert)
        elif risk_level == "medium":
            hotspots["medium_risk"].append(alert)
        else:
            hotspots["low_risk"].append(alert)
    
    return {
        "hotspots": hotspots,
        "summary": {
            "high_risk_count": len(hotspots["high_risk"]),
            "medium_risk_count": len(hotspots["medium_risk"]),
            "low_risk_count": len(hotspots["low_risk"]),
            "total_areas": len(alerts)
        }
    }


# ============================================================================
# DATA INTEGRITY VALIDATION
# ============================================================================

@router.post("/validate-measurement")
async def validate_measurement(
    body: OPTPlusCalculationRequest,
    user: User = Depends(get_current_user)
):
    """
    Validate measurement data against OPT Plus data integrity rules
    ⚠️ NOTE: OPT Plus features are disabled for admin and superadmin roles.
    
    Rules:
    - Overlapping Edge Logic: Scores exactly -2.00 are NORMAL
    - Drop-Day Age Factor: Discard days in age computation
    - Physical Adjustment Constants: +0.7 cm / -0.7 cm adjustments
    - Clinical Override Clause: Edema → SAM emergency
    """
    errors = []
    warnings = []
    
    # Age validation
    age_months = calculate_age_in_months(body.birth_date, body.measurement_date)
    if age_months > 71:
        errors.append(f"Age {age_months} months exceeds OPT Plus range (0-71 months)")
    
    # Weight validation
    if body.weight_kg < 1 or body.weight_kg > 50:
        warnings.append(f"Weight {body.weight_kg} kg seems unusual. Please verify.")
    
    # Height validation
    if body.height_cm < 40 or body.height_cm > 130:
        warnings.append(f"Height {body.height_cm} cm seems unusual. Please verify.")
    
    # Edema check
    if body.has_bilateral_edema:
        warnings.append("Bilateral edema present - will override Z-score calculations per OPT Plus guidelines")
    
    # Position check
    correct_position = "lying_down" if age_months < 24 else "standing"
    if body.measurement_position != correct_position:
        warnings.append(
            f"Measurement position '{body.measurement_position}' may not be optimal for age {age_months} months. "
            f"Recommended: '{correct_position}'. Adjustment of ±0.7 cm will be applied."
        )
    
    return {
        "valid": len(errors) == 0,
        "errors": errors,
        "warnings": warnings,
        "age_months": age_months,
        "recommended_position": correct_position
    }


# ============================================================================
# MODULE 6: BATCH IMPORT TO ASSESSMENT RECORDS
# ============================================================================

@router.get("/report")
async def get_opt_plus_report(
    year: int = Query(None, description="Filter data by year"),
    month: int = Query(None, description="Filter data by month (1-12)"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Get OPT Plus report data for dashboard
    ⚠️ NOTE: OPT Plus features are disabled for admin and superadmin roles.
    
    Returns aggregated child nutrition data with KPIs and summaries
    Accessible to all authenticated users
    """
    
    # Use current year/month if not specified
    if year is None:
        year = date.today().year
    if month is None:
        month = date.today().month
    
    # Calculate date range
    from calendar import monthrange
    
    start_date = datetime(year, month, 1).date()
    _, last_day = monthrange(year, month)
    end_date = datetime(year, month, last_day).date()
    
    try:
        # Get all active children aged 0-59 months with measurements in the period
        measurements_query = (
            select(Measurement)
            .options(selectinload(Measurement.child).selectinload(Child.barangay))
            .join(Child)
            .where(
                Child.is_active.is_(True),
                Measurement.age_in_months.between(0, 59),
                Measurement.measurement_date.between(start_date, end_date)
            )
        )
        
        measurements_result = await db.scalars(measurements_query)
        measurements = measurements_result.all()
        
        # Get barangay data for header
        barangays_result = await db.scalars(select(Barangay))
        barangays = barangays_result.all()
        
        # Get total population count
        total_population = sum([b.population_count or 0 for b in barangays]) or 10000
        
        # Initialize counters
        children_0_59_months = 0
        total_wfa = 0
        total_hfa = 0
        total_wflh = 0
        
        undernutrition_0_59 = 0
        overweight_0_59 = 0
        undernutrition_0_23 = 0
        overweight_0_23 = 0
        
        children_by_age_group = {
            "0-5": 0,
            "6-11": 0,
            "12-23": 0,
            "24-35": 0,
            "36-47": 0,
            "48-59": 0,
        }
        
        gender_breakdown = {
            "male": 0,
            "female": 0
        }
        
        # Process measurements
        for measurement in measurements:
            child = measurement.child
            age = measurement.age_in_months
            
            children_0_59_months += 1
            
            # Count by indicator
            if measurement.waz_status is not None:
                total_wfa += 1
            if measurement.haz_status is not None:
                total_hfa += 1
            if measurement.whz_status is not None:
                total_wflh += 1
            
            # Determine age group
            if age <= 5:
                age_group = "0-5"
            elif age <= 11:
                age_group = "6-11"
            elif age <= 23:
                age_group = "12-23"
            elif age <= 35:
                age_group = "24-35"
            elif age <= 47:
                age_group = "36-47"
            else:
                age_group = "48-59"
            
            children_by_age_group[age_group] += 1
            
            # Count gender
            if child and child.sex:
                if str(child.sex).upper() == "M":
                    gender_breakdown["male"] += 1
                else:
                    gender_breakdown["female"] += 1
            
            # Count nutritional issues
            # Using WAZ (Weight-for-Age) as primary indicator for undernutrition
            is_undernutrition = False
            is_overweight = False
            
            if measurement.waz_status in [WazStatus.underweight, WazStatus.severely_underweight]:
                is_undernutrition = True
            elif measurement.waz_status == WazStatus.overweight:
                is_overweight = True
            
            # Fallback to HAZ and WHZ if WAZ not available
            if measurement.haz_status in [HazStatus.stunted, HazStatus.severely_stunted]:
                is_undernutrition = True
            if measurement.whz_status in [WhzStatus.wasted, WhzStatus.severely_wasted]:
                is_undernutrition = True
            
            if is_undernutrition:
                undernutrition_0_59 += 1
                if age <= 23:
                    undernutrition_0_23 += 1
            
            if is_overweight:
                overweight_0_59 += 1
                if age <= 23:
                    overweight_0_23 += 1
        
        # Build age group data array
        age_group_data = []
        for age_group in ["0-5", "6-11", "12-23", "24-35", "36-47", "48-59"]:
            age_group_data.append({
                "age_group": age_group,
                "count": children_by_age_group[age_group]
            })
        
        # Get location info - get from Cabadbaran City configuration
        # Assuming Cabadbaran City is the primary municipality
        municipality = "Cabadbaran City"
        region = "Region XIII (CARAGA)"
        province = "Agusan del Norte"
        psgc = "053812000"  # Cabadbaran City PSGC code
        
        # Calculate coverage percentage
        # Coverage = (children measured / expected children 0-59m in area) * 100
        # Expected children = 12% of total population (demographic standard)
        expected_children_0_59 = max(1, total_population * 0.12)
        
        if children_0_59_months > 0:
            coverage_percentage = (children_0_59_months / expected_children_0_59) * 100
            coverage_percentage = min(100, coverage_percentage)  # Cap at 100%
        else:
            coverage_percentage = 0.0
        
        return {
            "province": province,
            "region": region,
            "municipality": municipality,
            "psgc": psgc,
            "total_population": total_population,
            "children_0_59_months": children_0_59_months,
            "coverage_percentage": coverage_percentage,
            "total_wfa": total_wfa,
            "total_hfa": total_hfa,
            "total_wflh": total_wflh,
            "age_group_data": age_group_data,
            "gender_breakdown": gender_breakdown,
            "summary": {
                "undernutrition_0_59": undernutrition_0_59,
                "overweight_0_59": overweight_0_59,
                "undernutrition_0_23": undernutrition_0_23,
                "overweight_0_23": overweight_0_23
            },
            "period": {
                "year": year,
                "month": month,
                "start_date": str(start_date),
                "end_date": str(end_date)
            }
        }
        
    except Exception as e:
        import traceback
        error_msg = f"Error generating OPT Plus report: {str(e)}"
        print(error_msg)
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=error_msg)


@router.post("/import-measurements")
async def import_measurements(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user)
):
    """
    Import children measurements from Excel file and save to Assessment Records.
    
    ** ROLE-BASED IMPORT **
    
    ADMIN (Individual Child Data):
    - Imports detailed measurements for specific named children
    - Children are auto-created if they don't exist (requires columns 10-15)
    - All children data is linked to actual child records
    
    SUPERADMIN (Aggregate Tally Data):
    - Imports aggregate/anonymous measurement data for reporting
    - Child names are optional (system generates anonymous entries)
    - Can provide age_months (column 16) instead of birth_date
    - Requires barangay_id and purok_id for data organization
    
    ** REQUIRED COLUMNS FOR ADMIN **
    - Column 1: full_name (required)
    - Column 2: birth_date (required, YYYY-MM-DD)
    - Column 3: sex (required, M/F)
    - Column 4: measurement_date (required, YYYY-MM-DD)
    - Column 5: weight_kg (required, decimal)
    - Column 6: height_cm (required, decimal)
    - Column 7: muac_cm (optional)
    - Column 8: notes (optional)
    - Column 9: child_id (optional, UUID)
    
    AUTO-CREATE (for admin when child not found):
    - Column 10: address
    - Column 11: latitude (decimal)
    - Column 12: longitude (decimal)
    - Column 13: guardian_name
    - Column 14: purok_id (UUID)
    - Column 15: barangay_id (UUID)
    
    ** REQUIRED COLUMNS FOR SUPERADMIN **
    - Column 1: full_name (optional, auto-generated if blank)
    - Column 2: birth_date (optional if age_months provided)
    - Column 3: sex (required, M/F)
    - Column 4: measurement_date (required, YYYY-MM-DD)
    - Column 5: weight_kg (required, decimal)
    - Column 6: height_cm (required, decimal)
    - Column 7: muac_cm (optional)
    - Column 8: notes (optional)
    - Column 14: purok_id (required, UUID)
    - Column 15: barangay_id (required, UUID)
    - Column 16: age_months (optional, use if birth_date not provided)
    
    Returns summary of imported records and any errors.
    """
    try:
        # Read Excel file
        contents = await file.read()
        wb = load_workbook(BytesIO(contents))
        ws = wb.active
        
        imported_count = 0
        error_count = 0
        errors = []
        imported_records = []
        
        # Get child records for faster lookup
        children_stmt = select(Child)
        children_result = await db.scalars(children_stmt)
        children_map = {str(child.id): child for child in children_result}
        
        # Detect e-OPT PLUS format by looking for the header row
        is_eopt_format = False
        data_start_row = 2  # Default for standard format
        
        # Scan first 20 rows to find "Child Seq" header
        for row_num in range(1, min(25, ws.max_row + 1)):
            row = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))[0] if ws.max_row >= row_num else []
            if row and len(row) > 0:
                # Check if first cell contains "Child Seq"
                first_cell = str(row[0]) if row[0] else ""
                if "Child Seq" in first_cell or "child seq" in first_cell.lower():
                    is_eopt_format = True
                    data_start_row = row_num + 1  # Data starts right after header
                    break
        
        print(f"DEBUG: Format detected: {'e-OPT PLUS' if is_eopt_format else 'Standard'}, Data starts at row {data_start_row}")
        
        # Process each row
        row_count = 0
        for idx, row in enumerate(ws.iter_rows(min_row=data_start_row, values_only=True), start=data_start_row):
            try:
                row_count += 1
                
                # Skip completely empty rows
                if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                    print(f"DEBUG Row {idx}: Skipped - completely empty")
                    continue
                
                # Extract values based on format
                if is_eopt_format:
                    # e-OPT PLUS format - exact columns from screenshot
                    full_name = row[3] if len(row) > 3 and row[3] else None  # Column D
                    mother_name = row[2] if len(row) > 2 and row[2] else None  # Column C
                    address_location = row[1] if len(row) > 1 and row[1] else None  # Column B
                    is_ip = row[4] if len(row) > 4 and row[4] else None  # Column E
                    sex_val = row[5] if len(row) > 5 and row[5] else None  # Column F
                    birth_date_val = row[6] if len(row) > 6 and row[6] else None  # Column G
                    measurement_date_val = row[7] if len(row) > 7 and row[7] else None  # Column H
                    weight_kg = row[8] if len(row) > 8 and row[8] is not None else None  # Column I
                    height_cm = row[9] if len(row) > 9 and row[9] is not None else None  # Column J
                    age_months_val = row[10] if len(row) > 10 and row[10] is not None else None  # Column K
                    
                    print(f"DEBUG Row {idx}: name={full_name}, sex={sex_val}, dob={birth_date_val}, meas_date={measurement_date_val}, weight={weight_kg}, height={height_cm}")
                    
                    muac_cm = None
                    notes = f"Location: {address_location}, IP: {is_ip}" if address_location or is_ip else None
                    child_id_val = None
                    
                    # Skip if no child name (likely a header or summary row)
                    if not full_name or str(full_name).strip() in ["", "(Surname, First Name)"]:
                        print(f"DEBUG Row {idx}: Skipped - no valid child name")
                        continue
                    
                else:
                    # Standard format
                    full_name = row[0] if len(row) > 0 and row[0] else None
                    birth_date_val = row[1] if len(row) > 1 and row[1] else None
                    sex_val = row[2] if len(row) > 2 and row[2] else None
                    measurement_date_val = row[3] if len(row) > 3 and row[3] else None
                    weight_kg = row[4] if len(row) > 4 and row[4] is not None else None
                    height_cm = row[5] if len(row) > 5 and row[5] is not None else None
                    muac_cm = row[6] if len(row) > 6 and row[6] else None
                    notes = row[7] if len(row) > 7 and row[7] else None
                    child_id_val = row[8] if len(row) > 8 and row[8] else None
                    age_months_val = row[15] if len(row) > 15 and row[15] else None
                    mother_name = row[12] if len(row) > 12 and row[12] else None
                    address_location = row[9] if len(row) > 9 and row[9] else None
                
                # Convert date values FIRST - Excel dates come as datetime objects
                if birth_date_val:
                    try:
                        if isinstance(birth_date_val, str):
                            birth_date = datetime.strptime(birth_date_val, "%Y-%m-%d").date()
                        elif hasattr(birth_date_val, 'date'):
                            # It's a datetime object
                            birth_date = birth_date_val.date()
                        else:
                            # Assume it's already a date object
                            birth_date = birth_date_val
                    except Exception as e:
                        errors.append(f"Row {idx}: Invalid birth date format in col {'G' if is_eopt_format else 'B'}: {birth_date_val}")
                        error_count += 1
                        continue
                else:
                    birth_date = None
                
                if measurement_date_val:
                    try:
                        if isinstance(measurement_date_val, str):
                            measurement_date = datetime.strptime(measurement_date_val, "%Y-%m-%d").date()
                        elif hasattr(measurement_date_val, 'date'):
                            # It's a datetime object
                            measurement_date = measurement_date_val.date()
                        else:
                            # Assume it's already a date object
                            measurement_date = measurement_date_val
                    except Exception as e:
                        errors.append(f"Row {idx}: Invalid measurement date format in col {'H' if is_eopt_format else 'D'}: {measurement_date_val}")
                        error_count += 1
                        continue
                else:
                    measurement_date = None
                
                # SUPERADMIN: Can import without child names (aggregate data)
                # ADMIN: Must provide child names (individual data)
                if user.role.value == "super_admin":
                    # For superadmin, generate anonymous child name if not provided
                    if not full_name:
                        full_name = f"Anonymous-{measurement_date_val}-{idx}"
                    # Birth date can be calculated from age if provided
                    if not birth_date and age_months_val and measurement_date:
                        from dateutil.relativedelta import relativedelta
                        birth_date = measurement_date - relativedelta(months=int(age_months_val))
                
                # Validate required fields based on role
                if user.role.value == "admin":
                    # Admin must provide complete child information
                    if not full_name or not birth_date or not sex_val:
                        missing_fields = []
                        if not full_name: missing_fields.append(f"name (col {'D' if is_eopt_format else 'A'}): '{full_name}'")
                        if not birth_date: missing_fields.append(f"birth_date (col {'G' if is_eopt_format else 'B'}): '{birth_date_val}' -> '{birth_date}'")
                        if not sex_val: missing_fields.append(f"sex (col {'F' if is_eopt_format else 'C'}): '{sex_val}'")
                        error_msg = f"Row {idx}: Missing required fields: {', '.join(missing_fields)}"
                        errors.append(error_msg)
                        print(f"DEBUG ERROR: {error_msg}")
                        print(f"  Raw data: name={full_name!r}, birth_val={birth_date_val!r}, birth_date={birth_date!r}, sex={sex_val!r}")
                        error_count += 1
                        continue
                else:
                    # Superadmin needs sex and either birth_date or age_months
                    if not sex_val:
                        errors.append(f"Row {idx}: Missing required field (sex)")
                        error_count += 1
                        continue
                    if not birth_date and not age_months_val:
                        errors.append(f"Row {idx}: Must provide either birth_date (column 2) or age_months (column 16)")
                        error_count += 1
                        continue
                
                if not measurement_date or weight_kg is None or height_cm is None:
                    missing_fields = []
                    if not measurement_date: missing_fields.append(f"measurement_date (col {'H' if is_eopt_format else 'D'}): '{measurement_date_val}' -> '{measurement_date}'")
                    if weight_kg is None: missing_fields.append(f"weight_kg (col {'I' if is_eopt_format else 'E'}): '{weight_kg}'")
                    if height_cm is None: missing_fields.append(f"height_cm (col {'J' if is_eopt_format else 'F'}): '{height_cm}'")
                    error_msg = f"Row {idx}: Missing measurement data: {', '.join(missing_fields)}"
                    errors.append(error_msg)
                    print(f"DEBUG ERROR: {error_msg}")
                    error_count += 1
                    continue
                
                # Normalize sex
                sex_normalized = "M" if str(sex_val).upper() in ["M", "MALE"] else "F"
                sex_name = "male" if sex_normalized == "M" else "female"
                
                # Find or create child based on role
                child = None
                
                if user.role.value == "super_admin":
                    # SUPERADMIN: Auto-create anonymous children for aggregate data
                    # Don't search for existing children, just create new ones for tally
                    
                    # Get location data - use defaults if not provided
                    address_val = row[9] if len(row) > 9 and row[9] else "Aggregate Data Import"
                    latitude_val = row[10] if len(row) > 10 and row[10] else 8.9483  # Default coordinates
                    longitude_val = row[11] if len(row) > 11 and row[11] else 125.5282
                    guardian_name_val = row[12] if len(row) > 12 and row[12] else "N/A"
                    purok_id_val = row[13] if len(row) > 13 and row[13] else None
                    barangay_id_val = row[14] if len(row) > 14 and row[14] else None
                    
                    # For superadmin, barangay and purok are required for data organization
                    if not barangay_id_val or not purok_id_val:
                        errors.append(f"Row {idx}: Barangay ID (column 15) and Purok ID (column 14) are required for aggregate data")
                        error_count += 1
                        continue
                    
                    try:
                        # Create anonymous child for aggregate data
                        new_child = Child(
                            full_name=full_name,  # Will be "Anonymous-date-rownum"
                            birth_date=birth_date if birth_date_val else measurement_date,  # Use measurement date if birth date not provided
                            sex=Sex.male if sex_normalized == "M" else Sex.female,
                            guardian_name=str(guardian_name_val),
                            purok_id=uuid.UUID(str(purok_id_val)),
                            barangay_id=uuid.UUID(str(barangay_id_val)),
                            latitude=float(latitude_val),
                            longitude=float(longitude_val),
                            contact_number=None,
                            household_id=None,
                            is_active=True
                        )
                        db.add(new_child)
                        await db.flush()
                        child = new_child
                        children_map[str(child.id)] = child
                    except Exception as create_error:
                        errors.append(f"Row {idx}: Failed to create aggregate data record: {str(create_error)}")
                        error_count += 1
                        continue
                        
                elif user.role.value == "admin":
                    # ADMIN: Search for existing children or auto-create with full details
                    if child_id_val:
                        child_id_str = str(child_id_val)
                        child = children_map.get(child_id_str)
                    
                    if not child:
                        # Try to find by name (case-insensitive)
                        name_search_stmt = select(Child).where(
                            Child.full_name.ilike(f"%{full_name}%")
                        )
                        child = (await db.scalars(name_search_stmt)).first()
                    
                    # If child still not found, auto-create with provided data
                    if not child:
                        # Get additional columns for child creation if available
                        if is_eopt_format:
                            # Use data from e-OPT format
                            guardian_name_val = mother_name if mother_name else full_name
                            address_val = address_location if address_location else "Unknown"
                            # e-OPT format doesn't have location columns, use defaults
                            latitude_val = 8.9483  # Default coords
                            longitude_val = 125.5282
                            purok_id_val = None
                            barangay_id_val = None
                        else:
                            address_val = row[9] if len(row) > 9 and row[9] else None
                            guardian_name_val = row[12] if len(row) > 12 and row[12] else full_name
                            latitude_val = row[10] if len(row) > 10 and row[10] else 8.9483  # Default coords
                            longitude_val = row[11] if len(row) > 11 and row[11] else 125.5282
                            purok_id_val = row[13] if len(row) > 13 and row[13] else None
                            barangay_id_val = row[14] if len(row) > 14 and row[14] else None
                        
                        # For admin, use their assigned barangay if not provided
                        if not barangay_id_val and user.barangay_id:
                            barangay_id_val = user.barangay_id
                            print(f"DEBUG Row {idx}: Using admin's barangay: {barangay_id_val}")
                        
                        # If still no barangay, we need to get one from system
                        if not barangay_id_val:
                            # Get first barangay as fallback
                            brgy_stmt = select(Barangay).limit(1)
                            first_brgy = (await db.scalars(brgy_stmt)).first()
                            if first_brgy:
                                barangay_id_val = first_brgy.id
                                print(f"DEBUG Row {idx}: Using first available barangay: {barangay_id_val}")
                        
                        # Get first purok of the barangay if not provided
                        if not purok_id_val and barangay_id_val:
                            purok_stmt = select(Purok).where(Purok.barangay_id == barangay_id_val).limit(1)
                            first_purok = (await db.scalars(purok_stmt)).first()
                            if first_purok:
                                purok_id_val = first_purok.id
                                print(f"DEBUG Row {idx}: Using first purok of barangay: {purok_id_val}")
                        
                        # Check if we have minimal required data
                        if not barangay_id_val or not purok_id_val:
                            error_msg = f"Row {idx}: Child '{full_name}' not found and cannot auto-create (no barangay/purok available in system)"
                            errors.append(error_msg)
                            print(f"DEBUG ERROR: {error_msg}")
                            print(f"  barangay_id_val={barangay_id_val}, purok_id_val={purok_id_val}")
                            print(f"  user.barangay_id={user.barangay_id if hasattr(user, 'barangay_id') else 'N/A'}")
                            error_count += 1
                            continue
                        
                        try:
                            # Create new child with provided data
                            print(f"DEBUG Row {idx}: Creating child '{full_name}' with barangay={barangay_id_val}, purok={purok_id_val}")
                            new_child = Child(
                                full_name=full_name,
                                birth_date=birth_date,
                                sex=Sex.male if sex_normalized == "M" else Sex.female,
                                guardian_name=str(guardian_name_val),
                                purok_id=uuid.UUID(str(purok_id_val)),
                                barangay_id=uuid.UUID(str(barangay_id_val)),
                                latitude=float(latitude_val),
                                longitude=float(longitude_val),
                                contact_number=None,
                                household_id=None,
                                is_active=True
                            )
                            db.add(new_child)
                            await db.flush()
                            child = new_child
                            children_map[str(child.id)] = child
                            print(f"DEBUG Row {idx}: Child created successfully with ID: {child.id}")
                        except Exception as create_error:
                            import traceback
                            error_msg = f"Row {idx}: Failed to auto-create child '{full_name}': {str(create_error)}"
                            errors.append(error_msg)
                            print(f"DEBUG ERROR: {error_msg}")
                            print(f"  Traceback: {traceback.format_exc()}")
                            error_count += 1
                            continue
                
                # Calculate age in months
                age_months = calculate_age_in_months(birth_date, measurement_date)
                
                # Validate age is in OPT Plus range
                if age_months > 71:
                    errors.append(f"Row {idx}: Child age {age_months} months exceeds OPT Plus range (0-71)")
                    error_count += 1
                    continue
                
                # Adjust height for position (assume standing for > 24 months, lying for < 24)
                position = "lying_down" if age_months < 24 else "standing"
                adjusted_height = adjust_height_for_position(height_cm, age_months, position)
                
                # Calculate full assessment
                assessment = calculate_full_opt_plus_assessment(
                    sex=sex_name,
                    birth_date=birth_date,
                    measurement_date=measurement_date,
                    weight_kg=float(weight_kg),
                    height_cm=adjusted_height,
                    measurement_position=position,
                    has_bilateral_edema=False
                )
                
                # Create Measurement record
                measurement = Measurement(
                    child_id=child.id,
                    measured_by=user.id,
                    measurement_date=measurement_date,
                    age_in_months=age_months,
                    weight_kg=float(weight_kg),
                    height_cm=adjusted_height,
                    muac_cm=float(muac_cm) if muac_cm else None,
                    waz=assessment["waz"],
                    haz=assessment["haz"],
                    whz=assessment["whz"],
                    waz_status=assessment["waz_status"],
                    haz_status=assessment["haz_status"],
                    whz_status=assessment["whz_status"],
                    overall_status=assessment["overall_status"]
                )
                
                db.add(measurement)
                imported_count += 1
                
                # Track imported record for response
                imported_records.append({
                    "child_name": child.full_name,
                    "age_months": age_months,
                    "weight_kg": float(weight_kg),
                    "height_cm": adjusted_height,
                    "status": assessment["overall_status"]
                })
                
            except Exception as e:
                import traceback
                error_msg = f"Row {idx}: {str(e)}"
                errors.append(error_msg)
                print(f"DEBUG EXCEPTION: {error_msg}")
                print(f"  Exception type: {type(e).__name__}")
                print(f"  Traceback: {traceback.format_exc()}")
                error_count += 1
                continue
        
        # Commit all records
        await db.commit()
        
        print(f"DEBUG: Import complete - Processed {row_count} rows, Imported {imported_count}, Errors {error_count}")
        
        # Categorize errors for better debugging
        error_categories = {}
        for err in errors:
            if "Missing required fields" in err:
                error_categories.setdefault("Missing Data", []).append(err)
            elif "Invalid" in err and "date" in err:
                error_categories.setdefault("Date Errors", []).append(err)
            elif "not found" in err or "auto-create" in err:
                error_categories.setdefault("Child Lookup/Creation", []).append(err)
            elif "Missing measurement data" in err:
                error_categories.setdefault("Missing Measurements", []).append(err)
            elif "exceeds OPT Plus range" in err:
                error_categories.setdefault("Age Range", []).append(err)
            else:
                error_categories.setdefault("Other Errors", []).append(err)
        
        # Build categorized error summary for console
        print("\n=== ERROR SUMMARY ===")
        for category, category_errors in error_categories.items():
            print(f"{category}: {len(category_errors)} errors")
            if len(category_errors) <= 3:
                for err in category_errors:
                    print(f"  - {err}")
        
        return {
            "success": True,
            "imported": imported_count,
            "errors": error_count,
            "error_details": errors,  # Return all errors
            "error_categories": error_categories,  # Categorized for frontend
            "imported_records": imported_records,
            "message": f"Successfully imported {imported_count} measurements. Processed {row_count} rows total."
        }
        
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=400, detail=f"Import failed: {str(e)}")

