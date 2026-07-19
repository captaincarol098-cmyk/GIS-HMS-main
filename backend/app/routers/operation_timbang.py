"""
Operation Timbang Plus Router - INTEGRATED VERSION
Nutritional Assessment & Monitoring Tool - Child Weight & Height Tracking
Based on Philippine DOH Operation Timbang Plus program

🎯 SYSTEM INTEGRATION:
- Operation Timbang records now sync to main Children database
- Records appear in Children Monitoring automatically
- Data integrated with Reports, Analytics, and all other features
- WHO standards calculations applied
"""

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy import or_, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from datetime import datetime, date
from uuid import UUID
import pandas as pd
import io
from ..database import get_db
from ..middleware.rbac import get_current_user
from ..models import User, Child, Measurement, Barangay, Purok
from ..models.entities import Sex, WazStatus, HazStatus, WhzStatus, OverallStatus

router = APIRouter(prefix="/api/operation-timbang", tags=["operation-timbang"])

def calculate_age_in_months(date_of_birth: date) -> int:
    """Calculate age in months from date of birth"""
    today = date.today()
    months = (today.year - date_of_birth.year) * 12 + (today.month - date_of_birth.month)
    return max(0, months)


def calculate_nutritional_status(weight: float, height: float, age_months: int) -> dict:
    """
    Calculate nutritional status based on WHO/NCHS standards
    Maps to available OverallStatus enum values
    """
    
    # WHO/NCHS Reference Standards by age (in months)
    who_standards = {
        6: {"weight_10": 5.5, "weight_50": 7.3, "height_5": 62, "height_50": 66},
        12: {"weight_10": 7.0, "weight_50": 9.6, "height_5": 69, "height_50": 75},
        18: {"weight_10": 8.4, "weight_50": 11.5, "height_5": 76, "height_50": 82},
        24: {"weight_10": 9.6, "weight_50": 13.0, "height_5": 82, "height_50": 88},
        30: {"weight_10": 10.6, "weight_50": 14.3, "height_5": 88, "height_50": 94},
        36: {"weight_10": 11.5, "weight_50": 15.5, "height_5": 93, "height_50": 100},
        42: {"weight_10": 12.4, "weight_50": 16.7, "height_5": 98, "height_50": 105},
        48: {"weight_10": 13.2, "weight_50": 17.8, "height_5": 103, "height_50": 110},
        54: {"weight_10": 13.9, "weight_50": 18.8, "height_5": 108, "height_50": 115},
        60: {"weight_10": 14.6, "weight_50": 19.8, "height_5": 112, "height_50": 120},
    }
    
    # Get closest age reference
    closest_age = min(who_standards.keys(), key=lambda x: abs(x - age_months))
    reference = who_standards.get(closest_age, who_standards[60])
    
    # Calculate nutritional status indicators
    weight_for_age_status = WazStatus.normal
    height_for_age_status = HazStatus.normal
    weight_for_height_status = WhzStatus.normal
    
    # Weight-for-Age Assessment (WFA)
    if weight < reference["weight_10"] * 0.9:
        weight_for_age_status = WazStatus.underweight
    elif weight > reference["weight_50"] * 1.2:
        weight_for_age_status = WazStatus.overweight
    
    # Height-for-Age Assessment (HFA)
    if height < reference["height_5"] * 0.95:
        height_for_age_status = HazStatus.stunted
    
    # Weight-for-Height Assessment (WFH) - BMI-based
    height_m = height / 100
    if height_m > 0:
        bmi = weight / (height_m ** 2)
        ref_height_m = reference["height_50"] / 100
        ref_bmi_normal = reference["weight_50"] / (ref_height_m ** 2)
        
        if bmi < ref_bmi_normal * 0.85:
            weight_for_height_status = WhzStatus.wasted
        elif bmi > ref_bmi_normal * 1.15:
            weight_for_height_status = WhzStatus.overweight
        elif bmi > ref_bmi_normal * 1.25:
            weight_for_height_status = WhzStatus.obese
    
    # Determine overall nutritional status
    # Priority: Wasted > Stunted > Underweight > Overweight > Normal
    # Map to available OverallStatus enum values
    overall_status = OverallStatus.normal  # Default
    
    if weight_for_height_status == WhzStatus.wasted:
        overall_status = OverallStatus.severe_acute_malnutrition  # Wasted = SAM
    elif height_for_age_status == HazStatus.stunted:
        overall_status = OverallStatus.moderate_acute_malnutrition  # Stunted = MAM
    elif weight_for_age_status == WazStatus.underweight:
        overall_status = OverallStatus.moderate_acute_malnutrition  # Underweight = MAM
    elif weight_for_height_status == WhzStatus.overweight or weight_for_age_status == WazStatus.overweight:
        overall_status = OverallStatus.overweight
    
    return {
        "waz_status": weight_for_age_status,
        "haz_status": height_for_age_status,
        "whz_status": weight_for_height_status,
        "overall_status": overall_status,
    }


@router.get("")
@router.get("")
async def get_records(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    search: str = Query(""),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Get Operation Timbang Plus records from main database - filtered by user's barangay
    ⚠️ NOTE: Operation Timbang features are disabled for admin and superadmin roles.
    
    NOTE: Aligned with dashboard analytics - only shows target age group (0-59 months)
    To see all children regardless of age, see /all-records endpoint
    """
    try:
        # Check if user is admin
        if user.role.value != "admin":
            raise HTTPException(status_code=403, detail="Admin access required")
        
        from ..services.analytics import TARGET_AGE_MIN, TARGET_AGE_MAX
        
        # Query recent measurements with eager loading of child relationship
        # Filter to match dashboard analytics (0-59 months only)
        stmt = (
            select(Measurement)
            .options(selectinload(Measurement.child).selectinload(Child.purok))
            .join(Child)
            .where(
                Measurement.age_in_months >= TARGET_AGE_MIN,
                Measurement.age_in_months <= TARGET_AGE_MAX,
                Child.is_active.is_(True)
            )
            .order_by(Measurement.measurement_date.desc())
        )
        
        # Filter by user's barangay if user is admin (not super_admin)
        if user.role.value == "admin" and user.barangay_id:
            stmt = stmt.where(Child.barangay_id == user.barangay_id)
        
        if search:
            search_lower = search.lower()
            # Search is already joined with Child, just add where conditions
            stmt = stmt.where(
                or_(
                    Child.full_name.ilike(f"%{search}%"),
                    Child.guardian_name.ilike(f"%{search}%")
                )
            )
        
        # Get filtered measurements
        measurements = (await db.scalars(stmt.offset(skip).limit(limit))).all()
        
        # Get total count for pagination
        count_stmt = (
            select(func.count(Measurement.id))
            .join(Child)
            .where(
                Measurement.age_in_months >= TARGET_AGE_MIN,
                Measurement.age_in_months <= TARGET_AGE_MAX,
                Child.is_active.is_(True)
            )
        )
        if user.role.value == "admin" and user.barangay_id:
            count_stmt = count_stmt.where(Child.barangay_id == user.barangay_id)
        total = await db.scalar(count_stmt)
        
        result = []
        for m in measurements:
            child = m.child  # Use the eager-loaded relationship
            result.append({
                "id": str(m.id),
                "child_id": str(m.child_id),
                "child_name": child.full_name if child else "Unknown",
                "mother_name": child.guardian_name if child else "Unknown",
                "location": child.purok.name if child and child.purok else "Unknown",
                "sex": child.sex.value if child else "M",
                "date_of_birth": child.birth_date.isoformat() if child else "",
                "actual_date_visit": m.measurement_date.isoformat(),
                "weight": m.weight_kg,
                "height": m.height_cm,
                "age_in_months": m.age_in_months,
                "weight_for_age": m.waz_status.value,
                "height_for_age": m.haz_status.value,
                "weight_for_height": m.whz_status.value,
                "nutritional_status": m.overall_status.value,
                "created_at": m.created_at.isoformat(),
            })
        
        return {
            "data": result,
            "total": total,
            "skip": skip,
            "limit": limit,
            "note": "Showing children aged 0-59 months (0-5 years) - standard reporting group"
        }
    except Exception as e:
        import traceback
        print(f"Error in get_records: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/all-records")
async def get_all_records(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    search: str = Query(""),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Get ALL Operation Timbang Plus records regardless of age
    ⚠️ NOTE: Operation Timbang features are disabled for admin and superadmin roles.
    
    This shows ALL children with measurements, including those outside the 0-59 month target group
    Use this to see the complete picture before filtering
    """
    try:
        # Check if user is admin
        if user.role.value != "admin":
            raise HTTPException(status_code=403, detail="Admin access required")
        
        # Query all measurements with eager loading of child relationship
        stmt = (
            select(Measurement)
            .options(selectinload(Measurement.child).selectinload(Child.purok))
            .join(Child)
            .where(Child.is_active.is_(True))
            .order_by(Measurement.measurement_date.desc())
        )
        
        # Filter by user's barangay if user is admin
        if user.role.value == "admin" and user.barangay_id:
            stmt = stmt.where(Child.barangay_id == user.barangay_id)
        
        if search:
            search_lower = search.lower()
            stmt = stmt.where(
                or_(
                    Child.full_name.ilike(f"%{search}%"),
                    Child.guardian_name.ilike(f"%{search}%")
                )
            )
        
        # Get filtered measurements
        measurements = (await db.scalars(stmt.offset(skip).limit(limit))).all()
        
        # Get total count for pagination
        count_stmt = (
            select(func.count(Measurement.id))
            .join(Child)
            .where(Child.is_active.is_(True))
        )
        if user.role.value == "admin" and user.barangay_id:
            count_stmt = count_stmt.where(Child.barangay_id == user.barangay_id)
        total = await db.scalar(count_stmt)
        
        # Also get the target age count for reference
        from ..services.analytics import TARGET_AGE_MIN, TARGET_AGE_MAX
        target_count_stmt = (
            select(func.count(Measurement.id))
            .join(Child)
            .where(
                Child.is_active.is_(True),
                Measurement.age_in_months >= TARGET_AGE_MIN,
                Measurement.age_in_months <= TARGET_AGE_MAX
            )
        )
        if user.role.value == "admin" and user.barangay_id:
            target_count_stmt = target_count_stmt.where(Child.barangay_id == user.barangay_id)
        target_total = await db.scalar(target_count_stmt)
        
        result = []
        for m in measurements:
            child = m.child
            in_target_group = TARGET_AGE_MIN <= m.age_in_months <= TARGET_AGE_MAX
            result.append({
                "id": str(m.id),
                "child_id": str(m.child_id),
                "child_name": child.full_name if child else "Unknown",
                "mother_name": child.guardian_name if child else "Unknown",
                "location": child.purok.name if child and child.purok else "Unknown",
                "sex": child.sex.value if child else "M",
                "date_of_birth": child.birth_date.isoformat() if child else "",
                "actual_date_visit": m.measurement_date.isoformat(),
                "weight": m.weight_kg,
                "height": m.height_cm,
                "age_in_months": m.age_in_months,
                "in_target_group": in_target_group,
                "weight_for_age": m.waz_status.value,
                "height_for_age": m.haz_status.value,
                "weight_for_height": m.whz_status.value,
                "nutritional_status": m.overall_status.value,
                "created_at": m.created_at.isoformat(),
            })
        
        return {
            "data": result,
            "total": total,
            "target_age_total": target_total,
            "skip": skip,
            "limit": limit,
            "note": f"Showing ALL children. {target_total} are in target group (0-59 months), {total - (target_total or 0)} are outside target age"
        }
    except Exception as e:
        import traceback
        print(f"Error in get_all_records: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


@router.post("")
async def create_record(
    payload: dict,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Create new Operation Timbang record - integrates with Children database
    ⚠️ NOTE: Operation Timbang features are disabled for admin and superadmin roles.
    """
    try:
        if user.role.value != "admin":
            raise HTTPException(status_code=403, detail="Admin access required")
        
        # Admin users must be assigned to a barangay
        if not user.barangay_id:
            raise HTTPException(status_code=400, detail="User is not assigned to any barangay")
        
        # Validate required fields
        required_fields = ["child_name", "date_of_birth", "actual_date_visit", "weight", "height"]
        missing_fields = [f for f in required_fields if not payload.get(f)]
        if missing_fields:
            raise HTTPException(status_code=400, detail=f"Missing required fields: {', '.join(missing_fields)}")
        
        # Parse dates
        try:
            dob_str = payload.get("date_of_birth", "")
            visit_str = payload.get("actual_date_visit", "")
            
            # Handle both ISO string and date string formats
            if isinstance(dob_str, str):
                # Remove time component if present
                dob_str = dob_str.split("T")[0]
                dob = datetime.strptime(dob_str, "%Y-%m-%d").date()
            else:
                dob = dob_str if isinstance(dob_str, date) else date.today()
                
            if isinstance(visit_str, str):
                # Remove time component if present
                visit_str = visit_str.split("T")[0]
                visit_date = datetime.strptime(visit_str, "%Y-%m-%d").date()
            else:
                visit_date = visit_str if isinstance(visit_str, date) else date.today()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid date format: {str(e)}")
        
        # Calculate age and status
        age_months = calculate_age_in_months(dob)
        nutritional = calculate_nutritional_status(
            float(payload.get("weight", 0)),
            float(payload.get("height", 0)),
            age_months
        )
        
        # Find or create child
        child_name = payload.get("child_name", "").strip()
        
        # Convert sex from M/F to male/female
        sex_value = payload.get("sex", "M").lower()
        if sex_value == "m":
            sex_value = "male"
        elif sex_value == "f":
            sex_value = "female"
        
        existing_child = await db.scalar(
            select(Child).where(
                Child.full_name == child_name,
                Child.birth_date == dob,
                Child.barangay_id == user.barangay_id  # Ensure child is from same barangay
            )
        )
        
        if existing_child:
            child = existing_child
        else:
            # Get user's barangay and get or create a purok
            user_barangay = await db.get(Barangay, user.barangay_id)
            if not user_barangay:
                raise HTTPException(status_code=400, detail="User's barangay not found")
            
            # Get location/purok from form or use first purok in barangay
            location_str = payload.get("location", "").strip()
            
            # Try to find matching purok
            purok = None
            if location_str:
                purok = await db.scalar(
                    select(Purok).where(
                        Purok.barangay_id == user.barangay_id,
                        Purok.name.ilike(f"%{location_str}%")
                    )
                )
            
            # If no matching purok, get first purok in barangay
            if not purok:
                purok = await db.scalar(
                    select(Purok).where(
                        Purok.barangay_id == user.barangay_id
                    ).order_by(Purok.name)
                )
            
            if not purok:
                raise HTTPException(status_code=400, detail=f"No purok found in {user_barangay.name}")
            
            child = Child(
                full_name=child_name,
                birth_date=dob,
                sex=Sex(sex_value),
                guardian_name=payload.get("mother_name", "").strip(),
                contact_number="",
                purok_id=purok.id,
                barangay_id=user.barangay_id,  # Assign to user's barangay
                latitude=0.0,
                longitude=0.0,
                is_active=True
            )
            db.add(child)
            await db.flush()
        
        # Create measurement
        measurement = Measurement(
            child_id=child.id,
            measured_by=user.id,
            measurement_date=visit_date,
            age_in_months=age_months,
            weight_kg=float(payload.get("weight", 0)),
            height_cm=float(payload.get("height", 0)),
            waz=0.0,
            haz=0.0,
            whz=0.0,
            waz_status=nutritional["waz_status"],
            haz_status=nutritional["haz_status"],
            whz_status=nutritional["whz_status"],
            overall_status=nutritional["overall_status"]
        )
        db.add(measurement)
        await db.commit()
        
        return {
            "status": "success",
            "message": "Record created and synced to Children database",
            "data": {
                "id": str(measurement.id),
                "child_id": str(child.id),
                "child_name": child.full_name,
                "nutritional_status": nutritional["overall_status"].value,
            },
        }
    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        import traceback
        error_detail = f"{type(e).__name__}: {str(e)}"
        print(f"Error creating record: {error_detail}")
        print(traceback.format_exc())
        raise HTTPException(status_code=400, detail=error_detail)


@router.get("/{record_id}")
async def get_record(
    record_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Get a specific Operation Timbang record - verify barangay access
    ⚠️ NOTE: Operation Timbang features are disabled for admin and superadmin roles.
    """
    try:
        if user.role.value != "admin":
            raise HTTPException(status_code=403, detail="Admin access required")
        
        measurement = await db.get(Measurement, record_id)
        if not measurement:
            raise HTTPException(status_code=404, detail="Record not found")
        
        child = await db.get(Child, measurement.child_id)
        if not child:
            raise HTTPException(status_code=404, detail="Child record not found")
        
        # Check barangay access for admin users
        if user.role.value == "admin" and child.barangay_id != user.barangay_id:
            raise HTTPException(status_code=403, detail="Access denied - record not in your barangay")
        
        return {
            "id": str(measurement.id),
            "child_id": str(measurement.child_id),
            "child_name": child.full_name if child else "Unknown",
            "nutritional_status": measurement.overall_status.value,
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/{record_id}")
async def update_record(
    record_id: UUID,
    payload: dict,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Update Operation Timbang record - verify barangay access
    ⚠️ NOTE: Operation Timbang features are disabled for admin and superadmin roles.
    """
    try:
        if user.role.value != "admin":
            raise HTTPException(status_code=403, detail="Admin access required")
        
        measurement = await db.get(Measurement, record_id)
        if not measurement:
            raise HTTPException(status_code=404, detail="Record not found")
        
        child = await db.get(Child, measurement.child_id)
        if not child:
            raise HTTPException(status_code=404, detail="Child not found")
        
        # Check barangay access for admin users
        if user.role.value == "admin" and child.barangay_id != user.barangay_id:
            raise HTTPException(status_code=403, detail="Access denied - record not in your barangay")
        
        # Update child
        if "child_name" in payload:
            child.full_name = payload["child_name"]
        if "mother_name" in payload:
            child.guardian_name = payload["mother_name"]
        
        # Parse dates
        dob_str = payload.get("date_of_birth", str(child.birth_date))
        visit_str = payload.get("actual_date_visit", str(measurement.measurement_date))
        
        dob = datetime.fromisoformat(dob_str.replace("Z", "+00:00")).date() if isinstance(dob_str, str) else dob_str
        visit_date = datetime.fromisoformat(visit_str.replace("Z", "+00:00")).date() if isinstance(visit_str, str) else visit_str
        
        # Recalculate
        age_months = calculate_age_in_months(dob)
        nutritional = calculate_nutritional_status(
            payload.get("weight", measurement.weight_kg),
            payload.get("height", measurement.height_cm),
            age_months
        )
        
        # Update measurement
        measurement.weight_kg = payload.get("weight", measurement.weight_kg)
        measurement.height_cm = payload.get("height", measurement.height_cm)
        measurement.measurement_date = visit_date
        measurement.age_in_months = age_months
        measurement.waz_status = nutritional["waz_status"]
        measurement.haz_status = nutritional["haz_status"]
        measurement.whz_status = nutritional["whz_status"]
        measurement.overall_status = nutritional["overall_status"]
        
        await db.commit()
        
        return {
            "status": "success",
            "message": "Record updated",
            "data": {"id": str(measurement.id)},
        }
    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/{record_id}")
async def delete_record(
    record_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Delete Operation Timbang record - verify barangay access
    ⚠️ NOTE: Operation Timbang features are disabled for admin and superadmin roles.
    """
    try:
        if user.role.value != "admin":
            raise HTTPException(status_code=403, detail="Admin access required")
        
        measurement = await db.get(Measurement, record_id)
        if not measurement:
            raise HTTPException(status_code=404, detail="Record not found")
        
        # Get child to verify barangay access
        child = await db.get(Child, measurement.child_id)
        if not child:
            raise HTTPException(status_code=404, detail="Child record not found")
        
        # Check barangay access for admin users
        if user.role.value == "admin" and child.barangay_id != user.barangay_id:
            raise HTTPException(status_code=403, detail="Access denied - record not in your barangay")
        
        await db.delete(measurement)
        await db.commit()
        
        return {
            "status": "success",
            "message": "Record deleted successfully",
        }
    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/bulk/delete-all")
async def delete_all_records(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Delete ALL Operation Timbang records for user's barangay (admin) or entire system (superadmin)
    
    ⚠️ WARNING: Operation Timbang features are disabled for admin and superadmin roles.
    This endpoint is not accessible to these users.
    
    ⚠️ WARNING: This is a destructive operation that cannot be undone!
    
    - Admin: Deletes all measurements for their assigned barangay only
    - SuperAdmin: Deletes ALL measurements across all barangays
    """
    try:
        # Build query based on user role
        if user.role.value == "super_admin":
            # SuperAdmin: Delete ALL measurements in the system
            from sqlalchemy import delete as sql_delete
            
            stmt = sql_delete(Measurement)
            result = await db.execute(stmt)
            deleted_count = result.rowcount
            
        elif user.role.value == "admin":
            # Admin: Delete only measurements for children in their barangay
            if not user.barangay_id:
                raise HTTPException(status_code=400, detail="Admin user has no assigned barangay")
            
            # Get all children IDs in this barangay
            children_stmt = select(Child.id).where(Child.barangay_id == user.barangay_id)
            children_result = await db.execute(children_stmt)
            child_ids = [row[0] for row in children_result.all()]
            
            if not child_ids:
                return {
                    "status": "success",
                    "message": "No records found to delete",
                    "deleted_count": 0,
                }
            
            # Delete all measurements for these children
            from sqlalchemy import delete as sql_delete
            stmt = sql_delete(Measurement).where(Measurement.child_id.in_(child_ids))
            result = await db.execute(stmt)
            deleted_count = result.rowcount
            
        else:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        
        await db.commit()
        
        role_context = "all barangays" if user.role.value == "super_admin" else f"barangay '{user.barangay_name}'"
        
        return {
            "status": "success",
            "message": f"All Operation Timbang records deleted successfully for {role_context}",
            "deleted_count": deleted_count,
            "user_role": user.role.value,
            "barangay_id": str(user.barangay_id) if user.barangay_id else None,
        }
        
    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        print(f"[DELETE ALL ERROR] {type(e).__name__}: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to delete records: {str(e)}")


@router.get("/stats/summary", tags=["stats"])
async def get_stats(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Get Operation Timbang Plus statistics - filtered by user's barangay
    ⚠️ NOTE: Operation Timbang features are disabled for admin and superadmin roles.
    """
    try:
        if user.role.value != "admin":
            raise HTTPException(status_code=403, detail="Admin access required")
        
        # Build queries with barangay filter for admin users
        stmt_normal = select(func.count(Measurement.id)).join(Child).where(
            Measurement.overall_status == OverallStatus.normal
        )
        stmt_underweight = select(func.count(Measurement.id)).join(Child).where(
            Measurement.overall_status == OverallStatus.moderate_acute_malnutrition
        )
        stmt_wasted = select(func.count(Measurement.id)).join(Child).where(
            Measurement.overall_status == OverallStatus.severe_acute_malnutrition
        )
        stmt_total = select(func.count(Measurement.id)).join(Child)
        
        # Add barangay filter for admin users
        if user.role.value == "admin" and user.barangay_id:
            stmt_normal = stmt_normal.where(Child.barangay_id == user.barangay_id)
            stmt_underweight = stmt_underweight.where(Child.barangay_id == user.barangay_id)
            stmt_wasted = stmt_wasted.where(Child.barangay_id == user.barangay_id)
            stmt_total = stmt_total.where(Child.barangay_id == user.barangay_id)
        
        total = await db.scalar(stmt_total) or 0
        normal = await db.scalar(stmt_normal) or 0
        underweight = await db.scalar(stmt_underweight) or 0
        wasted = await db.scalar(stmt_wasted) or 0
        
        return {
            "total": total,
            "normal": normal,
            "underweight": underweight,
            "wasted": wasted,
        }
    except Exception as e:
        import traceback
        print(f"Error in get_stats: {str(e)}")
        print(traceback.format_exc())
        raise HTTPException(status_code=500, detail=str(e))


# ============================================================================
# SUPERADMIN ENDPOINTS - e-OPT PLUS Format
# ============================================================================

@router.get("/superadmin/debug-dates")
async def debug_measurement_dates(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Debug endpoint to check what years have measurement data
    """
    if user.role.value != "super_admin":
        raise HTTPException(status_code=403, detail="SuperAdmin access required")
    
    from sqlalchemy import func, extract
    
    # Get year distribution
    stmt = (
        select(
            extract('year', Measurement.measurement_date).label('year'),
            func.count(Measurement.id).label('count')
        )
        .group_by(extract('year', Measurement.measurement_date))
        .order_by(extract('year', Measurement.measurement_date).desc())
    )
    
    results = (await db.execute(stmt)).all()
    
    year_distribution = [
        {"year": int(row.year) if row.year else None, "count": row.count}
        for row in results
    ]
    
    # Get sample dates from 2026
    sample_2026_stmt = (
        select(Measurement.measurement_date, Child.full_name)
        .join(Child)
        .where(extract('year', Measurement.measurement_date) == 2026)
        .limit(10)
    )
    
    sample_2026 = (await db.execute(sample_2026_stmt)).all()
    sample_dates_2026 = [
        {"date": str(row.measurement_date), "child": row.full_name}
        for row in sample_2026
    ]
    
    return {
        "year_distribution": year_distribution,
        "sample_dates_2026": sample_dates_2026,
        "total_measurements": sum(row["count"] for row in year_distribution)
    }


@router.post("/superadmin/move-2026-to-2025")
async def move_2026_measurements_to_2025(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Move all measurements from 2026 to 2025
    This fixes Excel import issues where dates were set to 2026 instead of 2025
    """
    if user.role.value != "super_admin":
        raise HTTPException(status_code=403, detail="SuperAdmin access required")
    
    from sqlalchemy import func, extract, update
    from datetime import timedelta
    
    try:
        # Count measurements in 2026 before update
        count_stmt = (
            select(func.count(Measurement.id))
            .where(extract('year', Measurement.measurement_date) == 2026)
        )
        count_before = await db.scalar(count_stmt) or 0
        
        if count_before == 0:
            return {
                "success": True,
                "message": "No measurements in 2026 to move",
                "moved_count": 0
            }
        
        # Update measurements: subtract 1 year from measurement_date
        update_stmt = (
            update(Measurement)
            .where(extract('year', Measurement.measurement_date) == 2026)
            .values(measurement_date=Measurement.measurement_date - timedelta(days=365))
        )
        
        await db.execute(update_stmt)
        await db.commit()
        
        # Verify the update
        count_after = await db.scalar(count_stmt) or 0
        
        return {
            "success": True,
            "message": f"Successfully moved {count_before} measurements from 2026 to 2025",
            "moved_count": count_before,
            "remaining_in_2026": count_after
        }
        
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Error moving measurements: {str(e)}")


@router.get("/superadmin/summary")
async def get_superadmin_opt_summary(
    year: int = Query(None, description="Filter data by year"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Get city-wide Operation Timbang Plus summary for SuperAdmin
    Returns data in e-OPT PLUS Tool format with barangay-level breakdown
    Supports year filtering for historical data
    Defaults to current year if not specified
    """
    if user.role.value != "super_admin":
        raise HTTPException(status_code=403, detail="SuperAdmin access required")

    # Use current year if not specified
    if year is None:
        year = date.today().year
    
    # Calculate date range for filtering
    from datetime import datetime
    start_date = datetime(year, 1, 1).date()
    end_date = datetime(year, 12, 31).date()

    try:
        # Get all barangays
        barangays_result = await db.scalars(select(Barangay).order_by(Barangay.name))
        barangays = barangays_result.all()

        barangay_data = []
        
        for idx, barangay in enumerate(barangays, 1):
            # Initialize age group structure
            age_groups = {
                "0-5": {"normal": 0, "overweight": 0, "severely_underweight": 0, "underweight": 0, 
                        "severely_stunted": 0, "stunted": 0, "tall": 0, "obese": 0, 
                        "severely_wasted": 0, "wasted": 0},
                "6-11": {"normal": 0, "overweight": 0, "severely_underweight": 0, "underweight": 0,
                         "severely_stunted": 0, "stunted": 0, "tall": 0, "obese": 0,
                         "severely_wasted": 0, "wasted": 0},
                "12-23": {"normal": 0, "overweight": 0, "severely_underweight": 0, "underweight": 0,
                          "severely_stunted": 0, "stunted": 0, "tall": 0, "obese": 0,
                          "severely_wasted": 0, "wasted": 0},
                "24-35": {"normal": 0, "overweight": 0, "severely_underweight": 0, "underweight": 0,
                          "severely_stunted": 0, "stunted": 0, "tall": 0, "obese": 0,
                          "severely_wasted": 0, "wasted": 0},
                "36-47": {"normal": 0, "overweight": 0, "severely_underweight": 0, "underweight": 0,
                          "severely_stunted": 0, "stunted": 0, "tall": 0, "obese": 0,
                          "severely_wasted": 0, "wasted": 0},
                "48-59": {"normal": 0, "overweight": 0, "severely_underweight": 0, "underweight": 0,
                          "severely_stunted": 0, "stunted": 0, "tall": 0, "obese": 0,
                          "severely_wasted": 0, "wasted": 0},
            }
            
            # Get measurements for this barangay with child data
            # Filter by year if specified
            query = (
                select(Measurement)
                .options(selectinload(Measurement.child))
                .join(Child)
                .where(
                    Child.barangay_id == barangay.id,
                    Child.is_active.is_(True),
                    Measurement.age_in_months.between(0, 59)
                )
            )
            
            # Year filter: check measurement_date year
            # Apply year filter to only show measurements from the specified year
            query = query.where(
                Measurement.measurement_date.between(start_date, end_date)
            )
            
            # Debug logging
            print(f"[DEBUG] Filtering for year {year}: {start_date} to {end_date}")
            
            measurements_result = await db.scalars(query)
            measurements = measurements_result.all()
            
            # Debug logging
            if len(measurements) > 0:
                print(f"[DEBUG] Barangay {barangay.name}: Found {len(measurements)} measurements")
                # Show first measurement date as sample
                if measurements:
                    print(f"[DEBUG] Sample measurement date: {measurements[0].measurement_date}")
            
            # Process measurements and categorize by age group
            below_normal_count = 0
            above_normal_count = 0
            children_0_23 = 0
            below_normal_0_23 = 0
            
            # Track unique children and mothers for calculations
            unique_mothers = set()
            mothers_with_below_normal = set()
            mothers_with_above_normal = set()
            
            for measurement in measurements:
                age = measurement.age_in_months
                child = measurement.child
                
                # Track mother/guardian if exists
                if child and child.guardian_name and child.guardian_name.strip():
                    unique_mothers.add(child.guardian_name.strip())
                
                # Note: indigenous_group field not yet in database schema
                # Will be added in future migration when needed
                
                # Determine age group
                if age <= 5:
                    age_key = "0-5"
                elif age <= 11:
                    age_key = "6-11"
                elif age <= 23:
                    age_key = "12-23"
                    children_0_23 += 1
                elif age <= 35:
                    age_key = "24-35"
                elif age <= 47:
                    age_key = "36-47"
                else:
                    age_key = "48-59"
                
                # Track if this child has below/above normal status
                is_below_normal = False
                is_above_normal = False
                
                # Count by nutritional status using the INDIVIDUAL status enums
                # WazStatus (Weight-for-Age)
                if measurement.waz_status == WazStatus.severely_underweight:
                    age_groups[age_key]["severely_underweight"] += 1
                    below_normal_count += 1
                    is_below_normal = True
                    if age <= 23:
                        below_normal_0_23 += 1
                elif measurement.waz_status == WazStatus.underweight:
                    age_groups[age_key]["underweight"] += 1
                    below_normal_count += 1
                    is_below_normal = True
                    if age <= 23:
                        below_normal_0_23 += 1
                elif measurement.waz_status == WazStatus.overweight:
                    age_groups[age_key]["overweight"] += 1
                    above_normal_count += 1
                    is_above_normal = True
                elif measurement.waz_status == WazStatus.normal:
                    age_groups[age_key]["normal"] += 1
                
                # HazStatus (Height-for-Age)
                if measurement.haz_status == HazStatus.severely_stunted:
                    age_groups[age_key]["severely_stunted"] += 1
                    if not is_below_normal:
                        below_normal_count += 1
                        is_below_normal = True
                        if age <= 23:
                            below_normal_0_23 += 1
                elif measurement.haz_status == HazStatus.stunted:
                    age_groups[age_key]["stunted"] += 1
                    if not is_below_normal:
                        below_normal_count += 1
                        is_below_normal = True
                        if age <= 23:
                            below_normal_0_23 += 1
                elif measurement.haz_status == HazStatus.tall:
                    age_groups[age_key]["tall"] += 1
                
                # WhzStatus (Weight-for-Height)
                if measurement.whz_status == WhzStatus.severely_wasted:
                    age_groups[age_key]["severely_wasted"] += 1
                    if not is_below_normal:
                        below_normal_count += 1
                        is_below_normal = True
                        if age <= 23:
                            below_normal_0_23 += 1
                elif measurement.whz_status == WhzStatus.wasted:
                    age_groups[age_key]["wasted"] += 1
                    if not is_below_normal:
                        below_normal_count += 1
                        is_below_normal = True
                        if age <= 23:
                            below_normal_0_23 += 1
                elif measurement.whz_status == WhzStatus.obese:
                    age_groups[age_key]["obese"] += 1
                    if not is_above_normal:
                        above_normal_count += 1
                        is_above_normal = True
                
                # Track mothers with nutritional issues
                if child and child.guardian_name and child.guardian_name.strip():
                    mother_name = child.guardian_name.strip()
                    if is_below_normal:
                        mothers_with_below_normal.add(mother_name)
                    if is_above_normal:
                        mothers_with_above_normal.add(mother_name)
            
            measurement_count = len(measurements)
            
            barangay_data.append({
                "sequence": idx,
                "barangay_name": barangay.name,
                "barangay_id": str(barangay.id),
                "population": barangay.population_count or 0,
                "valid_wfa": measurement_count,
                "valid_hfa": measurement_count,
                "valid_wflh": measurement_count,
                "age_groups": age_groups,
                "indigenous_children": 0,  # Field not yet in database schema
                "estimated_preschoolers": int((barangay.population_count or 0) * 0.08),
                "measured_preschoolers": measurement_count,
                "below_normal": below_normal_count,
                "above_normal": above_normal_count,
                "children_0_23_months": children_0_23,
                "below_normal_0_23": below_normal_0_23,
                "total_mothers": len(unique_mothers),
                "mothers_with_below_normal": len(mothers_with_below_normal),
                "mothers_with_above_normal": len(mothers_with_above_normal),
            })

        # Calculate summary statistics
        total_measurements = sum([b["measured_preschoolers"] for b in barangay_data])
        total_below_normal = sum([b["below_normal"] for b in barangay_data])
        total_above_normal = sum([b["above_normal"] for b in barangay_data])
        barangays_with_opt = len([b for b in barangay_data if b["measured_preschoolers"] > 0])
        coverage_percentage = (barangays_with_opt / len(barangays) * 100) if len(barangays) > 0 else 0

        return {
            "year": year,
            "filter_applied": f"{start_date} to {end_date}",
            "summary": {
                "total_barangays": len(barangays),
                "barangays_with_opt": barangays_with_opt,
                "coverage_percentage": coverage_percentage,
                "total_children_measured": total_measurements,
                "total_valid_wfa": total_measurements,
                "total_valid_hfa": total_measurements,
                "total_valid_wflh": total_measurements,
                "total_below_normal": total_below_normal,
                "total_above_normal": total_above_normal,
                "indigenous_children": 0,  # Field not yet in database schema
            },
            "barangays": barangay_data,
        }

    except Exception as e:
        import traceback
        error_msg = f"Error in superadmin summary: {str(e)}"
        print(error_msg)
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=error_msg)


@router.get("/superadmin/comprehensive-report")
async def get_comprehensive_report(
    year: int = Query(None, description="Filter data by year"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Get comprehensive Operation Timbang Plus report for SuperAdmin
    Returns city-wide overall statistics and summary
    Defaults to 2025 if not specified
    """
    if user.role.value != "super_admin":
        raise HTTPException(status_code=403, detail="SuperAdmin access required")

    # Default to 2025 if not specified
    if year is None:
        year = 2025
    
    # Calculate date range for filtering
    from datetime import datetime
    start_date = datetime(year, 1, 1).date()
    end_date = datetime(year, 12, 31).date()

    try:
        # Get all measurements for the year
        measurements_query = (
            select(Measurement)
            .options(selectinload(Measurement.child).selectinload(Child.barangay))
            .join(Child)
            .where(
                Child.is_active.is_(True),
                Measurement.age_in_months.between(0, 59),
                Measurement.measurement_date.between(start_date, end_date)
            )
        )
        
        measurements_result = await db.scalars(measurements_query)
        measurements = measurements_result.all()
        
        # Get all barangays for population data
        barangays_query = select(Barangay)
        barangays_result = await db.scalars(barangays_query)
        barangays = barangays_result.all()
        
        # Initialize counters
        total_children = 0
        total_below_normal = 0
        total_above_normal = 0
        
        # Age group breakdown
        age_group_counts = {
            "0-5": {"total": 0, "below_normal": 0, "above_normal": 0},
            "6-11": {"total": 0, "below_normal": 0, "above_normal": 0},
            "12-23": {"total": 0, "below_normal": 0, "above_normal": 0},
            "24-35": {"total": 0, "below_normal": 0, "above_normal": 0},
            "36-47": {"total": 0, "below_normal": 0, "above_normal": 0},
            "48-59": {"total": 0, "below_normal": 0, "above_normal": 0},
        }
        
        # Nutritional status breakdown
        nutritional_status = {
            "severely_underweight": 0,
            "underweight": 0,
            "normal": 0,
            "overweight": 0,
            "severely_stunted": 0,
            "stunted": 0,
            "tall": 0,
            "severely_wasted": 0,
            "wasted": 0,
            "obese": 0,
        }
        
        # Gender breakdown
        gender_breakdown = {"male": 0, "female": 0}
        
        # Barangay-level summary
        barangay_summary = {}
        
        # Process each measurement
        for measurement in measurements:
            child = measurement.child
            age = measurement.age_in_months
            
            total_children += 1
            
            # Get age group key
            if age <= 5:
                age_key = "0-5"
            elif age <= 11:
                age_key = "6-11"
            elif age <= 23:
                age_key = "12-23"
            elif age <= 35:
                age_key = "24-35"
            elif age <= 47:
                age_key = "36-47"
            else:
                age_key = "48-59"
            
            age_group_counts[age_key]["total"] += 1
            
            # Track if below or above normal
            is_below_normal = False
            is_above_normal = False
            
            # Count WAZ status
            if measurement.waz_status == WazStatus.severely_underweight:
                nutritional_status["severely_underweight"] += 1
                is_below_normal = True
            elif measurement.waz_status == WazStatus.underweight:
                nutritional_status["underweight"] += 1
                is_below_normal = True
            elif measurement.waz_status == WazStatus.normal:
                nutritional_status["normal"] += 1
            elif measurement.waz_status == WazStatus.overweight:
                nutritional_status["overweight"] += 1
                is_above_normal = True
            
            # Count HAZ status
            if measurement.haz_status == HazStatus.severely_stunted:
                nutritional_status["severely_stunted"] += 1
                if not is_below_normal:
                    is_below_normal = True
            elif measurement.haz_status == HazStatus.stunted:
                nutritional_status["stunted"] += 1
                if not is_below_normal:
                    is_below_normal = True
            elif measurement.haz_status == HazStatus.tall:
                nutritional_status["tall"] += 1
            
            # Count WHZ status
            if measurement.whz_status == WhzStatus.severely_wasted:
                nutritional_status["severely_wasted"] += 1
                if not is_below_normal:
                    is_below_normal = True
            elif measurement.whz_status == WhzStatus.wasted:
                nutritional_status["wasted"] += 1
                if not is_below_normal:
                    is_below_normal = True
            elif measurement.whz_status == WhzStatus.obese:
                nutritional_status["obese"] += 1
                if not is_above_normal:
                    is_above_normal = True
            
            # Track below/above normal
            if is_below_normal:
                total_below_normal += 1
                age_group_counts[age_key]["below_normal"] += 1
            if is_above_normal:
                total_above_normal += 1
                age_group_counts[age_key]["above_normal"] += 1
            
            # Track gender
            if child and child.sex:
                if str(child.sex).upper() == "M":
                    gender_breakdown["male"] += 1
                else:
                    gender_breakdown["female"] += 1
            
            # Track by barangay
            if child and child.barangay:
                barangay_name = child.barangay.name
                if barangay_name not in barangay_summary:
                    barangay_summary[barangay_name] = {
                        "children_count": 0,
                        "below_normal": 0,
                        "above_normal": 0
                    }
                barangay_summary[barangay_name]["children_count"] += 1
                if is_below_normal:
                    barangay_summary[barangay_name]["below_normal"] += 1
                if is_above_normal:
                    barangay_summary[barangay_name]["above_normal"] += 1
        
        # Calculate total population from barangays
        total_population = sum([b.population_count or 0 for b in barangays])
        
        # Calculate percentages
        below_normal_percentage = (total_below_normal / total_children * 100) if total_children > 0 else 0
        above_normal_percentage = (total_above_normal / total_children * 100) if total_children > 0 else 0
        barangays_with_measurements = len([b for b in barangay_summary.values() if b["children_count"] > 0])
        
        # Sort barangays by children count (descending)
        sorted_barangays = sorted(
            [{"name": k, **v} for k, v in barangay_summary.items()],
            key=lambda x: x["children_count"],
            reverse=True
        )

        return {
            "year": year,
            "period": f"{start_date} to {end_date}",
            "timestamp": datetime.now().isoformat(),
            "summary": {
                "total_population": total_population,
                "total_children_measured": total_children,
                "total_barangays": len(barangays),
                "barangays_with_measurements": barangays_with_measurements,
                "total_below_normal": total_below_normal,
                "below_normal_percentage": round(below_normal_percentage, 2),
                "total_above_normal": total_above_normal,
                "above_normal_percentage": round(above_normal_percentage, 2),
                "total_normal": nutritional_status["normal"],
            },
            "age_groups": age_group_counts,
            "nutritional_status": nutritional_status,
            "gender_breakdown": gender_breakdown,
            "barangay_breakdown": sorted_barangays,
        }

    except Exception as e:
        import traceback
        error_msg = f"Error generating comprehensive report: {str(e)}"
        print(error_msg)
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=error_msg)




@router.get("/superadmin/opt-analytics")
async def get_opt_analytics(
    year: int = Query(None, description="Filter data by year"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Get detailed OPT Plus Analytics with nutritional status breakdowns and barangay rankings
    Returns comprehensive analytics data for the specified year
    Defaults to current year if not specified
    Accessible to: SuperAdmin (citywide), Admin (barangay-specific data)
    """
    # Default to current year
    if year is None:
        year = date.today().year
    
    # Calculate date range
    from datetime import datetime
    start_date = datetime(year, 1, 1).date()
    end_date = datetime(year, 12, 31).date()

    try:
        # Get all measurements for the year
        measurements_query = (
            select(Measurement)
            .options(selectinload(Measurement.child).selectinload(Child.barangay))
            .join(Child)
            .where(
                Child.is_active.is_(True),
                Measurement.age_in_months.between(0, 59),
                Measurement.measurement_date.between(start_date, end_date)
            )
        )
        
        # Filter by barangay if user is admin
        if user.role.value == "admin" and user.barangay_id:
            measurements_query = measurements_query.where(Child.barangay_id == user.barangay_id)
        
        measurements_result = await db.scalars(measurements_query)
        measurements = measurements_result.all()
        
        # Initialize counters
        total_children = len(measurements)
        total_below_normal = 0
        total_above_normal = 0
        total_normal = 0
        
        # Nutritional status breakdown
        wfa_breakdown = {
            "normal": {"boys": 0, "girls": 0, "total": 0, "percentage": 0},
            "overweight": {"boys": 0, "girls": 0, "total": 0, "percentage": 0},
            "underweight": {"boys": 0, "girls": 0, "total": 0, "percentage": 0},
            "severely_underweight": {"boys": 0, "girls": 0, "total": 0, "percentage": 0},
        }
        
        hfa_breakdown = {
            "normal": {"boys": 0, "girls": 0, "total": 0, "percentage": 0},
            "tall": {"boys": 0, "girls": 0, "total": 0, "percentage": 0},
            "stunted": {"boys": 0, "girls": 0, "total": 0, "percentage": 0},
            "severely_stunted": {"boys": 0, "girls": 0, "total": 0, "percentage": 0},
        }
        
        whz_breakdown = {
            "normal": {"boys": 0, "girls": 0, "total": 0, "percentage": 0},
            "overweight": {"boys": 0, "girls": 0, "total": 0, "percentage": 0},
            "obese": {"boys": 0, "girls": 0, "total": 0, "percentage": 0},
            "moderately_wasted": {"boys": 0, "girls": 0, "total": 0, "percentage": 0},
            "severely_wasted": {"boys": 0, "girls": 0, "total": 0, "percentage": 0},
        }
        
        # Barangay rankings
        barangay_data = {}
        
        # Process measurements
        for measurement in measurements:
            child = measurement.child
            is_boy = child and str(child.sex).upper() == "M"
            gender_key = "boys" if is_boy else "girls"
            
            # Count WFA status
            if measurement.waz_status == WazStatus.severely_underweight:
                wfa_breakdown["severely_underweight"][gender_key] += 1
                wfa_breakdown["severely_underweight"]["total"] += 1
                total_below_normal += 1
            elif measurement.waz_status == WazStatus.underweight:
                wfa_breakdown["underweight"][gender_key] += 1
                wfa_breakdown["underweight"]["total"] += 1
                total_below_normal += 1
            elif measurement.waz_status == WazStatus.normal:
                wfa_breakdown["normal"][gender_key] += 1
                wfa_breakdown["normal"]["total"] += 1
                total_normal += 1
            elif measurement.waz_status == WazStatus.overweight:
                wfa_breakdown["overweight"][gender_key] += 1
                wfa_breakdown["overweight"]["total"] += 1
                total_above_normal += 1
            
            # Count HFA status
            if measurement.haz_status == HazStatus.severely_stunted:
                hfa_breakdown["severely_stunted"][gender_key] += 1
                hfa_breakdown["severely_stunted"]["total"] += 1
            elif measurement.haz_status == HazStatus.stunted:
                hfa_breakdown["stunted"][gender_key] += 1
                hfa_breakdown["stunted"]["total"] += 1
            elif measurement.haz_status == HazStatus.normal:
                hfa_breakdown["normal"][gender_key] += 1
                hfa_breakdown["normal"]["total"] += 1
            elif measurement.haz_status == HazStatus.tall:
                hfa_breakdown["tall"][gender_key] += 1
                hfa_breakdown["tall"]["total"] += 1
            
            # Count WHZ status
            if measurement.whz_status == WhzStatus.severely_wasted:
                whz_breakdown["severely_wasted"][gender_key] += 1
                whz_breakdown["severely_wasted"]["total"] += 1
            elif measurement.whz_status == WhzStatus.wasted:
                whz_breakdown["moderately_wasted"][gender_key] += 1
                whz_breakdown["moderately_wasted"]["total"] += 1
            elif measurement.whz_status == WhzStatus.normal:
                whz_breakdown["normal"][gender_key] += 1
                whz_breakdown["normal"]["total"] += 1
            elif measurement.whz_status == WhzStatus.overweight:
                whz_breakdown["overweight"][gender_key] += 1
                whz_breakdown["overweight"]["total"] += 1
            elif measurement.whz_status == WhzStatus.obese:
                whz_breakdown["obese"][gender_key] += 1
                whz_breakdown["obese"]["total"] += 1
            
            # Barangay tracking
            if child and child.barangay:
                barangay_name = child.barangay.name
                if barangay_name not in barangay_data:
                    barangay_data[barangay_name] = {
                        "uw": 0, "st": 0, "ws": 0,
                        "uw_count": 0, "st_count": 0, "ws_count": 0,
                        "total": 0
                    }
                barangay_data[barangay_name]["total"] += 1
                
                if measurement.waz_status in [WazStatus.underweight, WazStatus.severely_underweight]:
                    barangay_data[barangay_name]["uw_count"] += 1
                if measurement.haz_status in [HazStatus.stunted, HazStatus.severely_stunted]:
                    barangay_data[barangay_name]["st_count"] += 1
                if measurement.whz_status in [WhzStatus.wasted, WhzStatus.severely_wasted]:
                    barangay_data[barangay_name]["ws_count"] += 1
        
        # Calculate percentages
        for status in wfa_breakdown:
            total = wfa_breakdown[status]["total"]
            wfa_breakdown[status]["percentage"] = (total / total_children * 100) if total_children > 0 else 0
        
        for status in hfa_breakdown:
            total = hfa_breakdown[status]["total"]
            hfa_breakdown[status]["percentage"] = (total / total_children * 100) if total_children > 0 else 0
        
        for status in whz_breakdown:
            total = whz_breakdown[status]["total"]
            whz_breakdown[status]["percentage"] = (total / total_children * 100) if total_children > 0 else 0
        
        # Create barangay rankings
        underweight_rankings = []
        stunting_rankings = []
        wasting_rankings = []
        
        for barangay_name, data in sorted(barangay_data.items(), key=lambda x: x[1]["total"], reverse=True):
            coverage = (data["total"] / total_children * 100) if total_children > 0 else 0
            
            if data["total"] > 0:
                uw_prevalence = (data["uw_count"] / data["total"]) * 100
                st_prevalence = (data["st_count"] / data["total"]) * 100
                ws_prevalence = (data["ws_count"] / data["total"]) * 100
            else:
                uw_prevalence = st_prevalence = ws_prevalence = 0
            
            underweight_rankings.append({
                "barangay": barangay_name,
                "coverage": coverage,
                "prevalence": uw_prevalence,
                "affected": data["uw_count"]
            })
            
            stunting_rankings.append({
                "barangay": barangay_name,
                "coverage": coverage,
                "prevalence": st_prevalence,
                "affected": data["st_count"]
            })
            
            wasting_rankings.append({
                "barangay": barangay_name,
                "coverage": coverage,
                "prevalence": ws_prevalence,
                "affected": data["ws_count"]
            })
        
        # Sort and add ranks
        underweight_rankings.sort(key=lambda x: x["prevalence"], reverse=True)
        stunting_rankings.sort(key=lambda x: x["prevalence"], reverse=True)
        wasting_rankings.sort(key=lambda x: x["prevalence"], reverse=True)
        
        for idx, item in enumerate(underweight_rankings, 1):
            item["rank"] = idx
        for idx, item in enumerate(stunting_rankings, 1):
            item["rank"] = idx
        for idx, item in enumerate(wasting_rankings, 1):
            item["rank"] = idx

        return {
            "year": year,
            "summary": {
                "total_children_measured": total_children,
                "total_below_normal": total_below_normal,
                "below_normal_percentage": (total_below_normal / total_children * 100) if total_children > 0 else 0,
                "total_above_normal": total_above_normal,
                "above_normal_percentage": (total_above_normal / total_children * 100) if total_children > 0 else 0,
                "total_normal": total_normal,
            },
            "nutritional_status_breakdown": {
                "wfa": wfa_breakdown,
                "hfa": hfa_breakdown,
                "whz": whz_breakdown,
            },
            "age_group_breakdown": {},  # Can add later if needed
            "barangay_rankings": {
                "underweight": underweight_rankings,
                "stunting": stunting_rankings,
                "wasting": wasting_rankings,
            }
        }

    except Exception as e:
        import traceback
        error_msg = f"Error generating OPT Plus analytics: {str(e)}"
        print(error_msg)
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=error_msg)


@router.get("/superadmin/export-excel")
async def export_opt_plus_excel(
    year: int = Query(None, description="Year to export data for"),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Export Operation Timbang Plus data to Excel in e-OPT PLUS format
    Matches the 68-column format from DOH e-OPT PLUS Tool
    Defaults to current year if not specified
    """
    if user.role.value != "super_admin":
        raise HTTPException(status_code=403, detail="SuperAdmin access required")

    # Use current year if not specified
    if year is None:
        year = date.today().year

    try:
        from fastapi.responses import StreamingResponse
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        # Get summary data (filtered by year)
        summary_data = await get_superadmin_opt_summary(year, db, user)
        barangays = summary_data["barangays"]
        summary = summary_data["summary"]

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "OPT Plus Summary"

        # Set column widths
        ws.column_dimensions['A'].width = 5   # Sequence
        ws.column_dimensions['B'].width = 20  # Barangay
        
        # Header Section
        ws['A1'] = "Revised March 2022"
        ws['A1'].font = Font(bold=True, size=10)
        
        ws['B1'] = "e-OPT PLUS Tool for Municipalities and Cities"
        ws['B1'].font = Font(bold=True, size=14)
        ws.merge_cells('B1:G1')
        
        # Year header on right side
        year_col = 68  # Last column
        ws.cell(row=1, column=year_col-1, value="OPT Plus Year:")
        ws.cell(row=1, column=year_col, value=year)
        ws.cell(row=1, column=year_col).font = Font(bold=True, size=12)
        
        # Municipality/City info
        ws['A2'] = "Municipality/City:"
        ws['B2'] = "CABADBARAN CITY"
        ws['B2'].font = Font(bold=True)
        ws['D2'] = "Province:"
        ws['E2'] = "AGUSAN DEL NORTE"
        ws['E2'].font = Font(bold=True)
        ws['G2'] = "Region:"
        ws['H2'] = "CARAGA (Region XIII)"
        ws['H2'].font = Font(bold=True)
        
        # Summary statistics row
        ws['A3'] = f"Total Barangays: {summary['total_barangays']}"
        ws['D3'] = f"Coverage: {summary['coverage_percentage']:.1f}%"
        ws['G3'] = f"Total Children: {summary['total_children_measured']}"
        ws['J3'] = f"Below Normal: {summary['total_below_normal']}"
        ws['M3'] = f"Above Normal: {summary['total_above_normal']}"
        
        # Empty row
        ws.append([])
        
        # MAIN HEADER ROW - Row 5
        current_row = 5
        col_idx = 1
        
        # Define header structure
        headers = [
            ("Sequence", 1, 3),
            ("Barangay", 1, 3),
            ("No. of Children w/ Valid WFA", 1, 3),
            ("No. of Children w/ Valid HFA", 1, 3),
            ("No. of Children w/ Valid WFL/H", 1, 3),
            ("Barangay Population", 1, 3),
        ]
        
        # Add age group headers (6 groups × 10 indicators)
        age_groups = [
            ("0-5 Months", "0099CC"),
            ("6-11 Months", "00CC99"),
            ("12-23 Months", "00CCCC"),
            ("24-35 Months", "0099FF"),
            ("36-47 Months", "6666FF"),
            ("48-59 Months", "9966FF")
        ]
        
        # Write basic headers
        for header, col_span, row_span in headers:
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True, size=9, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            if col_span > 1:
                ws.merge_cells(start_row=current_row, start_column=col_idx, 
                             end_row=current_row, end_column=col_idx + col_span - 1)
            if row_span > 1:
                ws.merge_cells(start_row=current_row, start_column=col_idx,
                             end_row=current_row + row_span - 1, end_column=col_idx)
            col_idx += 1
        
        # Add age group headers
        for age_group, color in age_groups:
            cell = ws.cell(row=current_row, column=col_idx, value=age_group)
            cell.font = Font(bold=True, size=9, color="FFFFFF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=current_row, start_column=col_idx,
                         end_row=current_row, end_column=col_idx + 9)
            col_idx += 10
        
        # Add remaining headers
        remaining_headers = [
            ("# IP Children 0-59m", "FF6600"),
            ("Estimated # of 0-59 mos PS", "666666"),
            ("Tally of PS with measurements", "666666"),
            ("0-59 mos w/ Below Normal NS", "CC0000"),
            ("0-59 mos w/ overwt or Ob", "FFCC00"),
            ("# Children 0-23 months old", "FF66CC"),
            ("# 0-23 mos w/ Below Normal NS", "CC0066"),
            ("# Mothers/Caregivers of 0-59m", "9966CC"),
            ("# M/Cs of 0-59m w/ Below N child", "CC3366"),
            ("# M/Cs of 0-59m w/ Overwt or Ob child", "FFAA00"),
        ]
        
        for header, color in remaining_headers:
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True, size=8, color="FFFFFF")
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.merge_cells(start_row=current_row, start_column=col_idx,
                         end_row=current_row + 2, end_column=col_idx)
            ws.column_dimensions[get_column_letter(col_idx)].width = 8
            col_idx += 1
        
        # SUBHEADER ROW - Nutritional Status Codes (Row 6 & 7)
        indicators = ["N", "OW", "SUW", "UW", "SS", "S", "T", "Ob", "SW", "W"]
        
        for age_idx in range(6):
            start_col = 7 + (age_idx * 10)
            for ind_idx, indicator in enumerate(indicators):
                cell = ws.cell(row=current_row + 1, column=start_col + ind_idx, value=indicator)
                cell.font = Font(bold=True, size=8)
                cell.alignment = Alignment(horizontal="center")
                ws.column_dimensions[get_column_letter(start_col + ind_idx)].width = 4
        
        # DATA ROWS - Starting from row 8
        data_start_row = 8
        for idx, b in enumerate(barangays):
            row = data_start_row + idx
            col = 1
            
            # Basic info
            ws.cell(row=row, column=col, value=b["sequence"])
            col += 1
            ws.cell(row=row, column=col, value=b["barangay_name"])
            col += 1
            ws.cell(row=row, column=col, value=b["valid_wfa"])
            col += 1
            ws.cell(row=row, column=col, value=b["valid_hfa"])
            col += 1
            ws.cell(row=row, column=col, value=b["valid_wflh"])
            col += 1
            ws.cell(row=row, column=col, value=b["population"])
            col += 1
            
            # Age group data (6 groups × 10 indicators)
            for age_group in ["0-5", "6-11", "12-23", "24-35", "36-47", "48-59"]:
                age_data = b["age_groups"].get(age_group, {})
                for indicator in ["normal", "overweight", "severely_underweight", "underweight",
                                "severely_stunted", "stunted", "tall", "obese",
                                "severely_wasted", "wasted"]:
                    ws.cell(row=row, column=col, value=age_data.get(indicator, 0))
                    col += 1
            
            # Summary columns
            ws.cell(row=row, column=col, value=b["indigenous_children"])
            col += 1
            ws.cell(row=row, column=col, value=b["estimated_preschoolers"])
            col += 1
            ws.cell(row=row, column=col, value=b["measured_preschoolers"])
            col += 1
            ws.cell(row=row, column=col, value=b["below_normal"])
            col += 1
            ws.cell(row=row, column=col, value=b["above_normal"])
            col += 1
            ws.cell(row=row, column=col, value=b["children_0_23_months"])
            col += 1
            ws.cell(row=row, column=col, value=b["below_normal_0_23"])
            col += 1
            ws.cell(row=row, column=col, value=b["total_mothers"])
            col += 1
            ws.cell(row=row, column=col, value=b["mothers_with_below_normal"])
            col += 1
            ws.cell(row=row, column=col, value=b["mothers_with_above_normal"])
            col += 1

        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=e-OPT_PLUS_{year}_Cabadbaran.xlsx"}
        )

    except Exception as e:
        import traceback
        print(f"Export error: {str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to export Excel: {str(e)}")


@router.post("/superadmin/import-excel")
async def import_opt_plus_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Import Operation Timbang Plus data from Excel
    Accepts e-OPT PLUS format Excel files
    """
    if user.role.value != "super_admin":
        raise HTTPException(status_code=403, detail="SuperAdmin access required")

    try:
        # Read Excel file
        contents = await file.read()
        
        # Try to read with pandas
        try:
            # Skip header rows and read data
            df = pd.read_excel(io.BytesIO(contents), skiprows=3)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid Excel file format: {str(e)}")
        
        # Validate required columns
        required_columns = ["Sequence", "Barangay"]
        missing = [col for col in required_columns if col not in df.columns]
        if missing:
            raise HTTPException(
                status_code=400, 
                detail=f"Missing required columns: {', '.join(missing)}"
            )
        
        imported_count = 0
        error_count = 0
        error_details = []
        
        # Process each row
        for idx, row in df.iterrows():
            try:
                barangay_name = str(row.get("Barangay", "")).strip()
                if not barangay_name or pd.isna(barangay_name):
                    continue
                
                # Find barangay in database
                barangay = await db.scalar(
                    select(Barangay).where(Barangay.name.ilike(f"%{barangay_name}%"))
                )
                
                if not barangay:
                    error_details.append(f"Row {idx + 5}: Barangay '{barangay_name}' not found in database")
                    error_count += 1
                    continue
                
                # Note: This is a summary import - individual child records would need more detailed data
                # For now, we just validate the barangay exists
                imported_count += 1
                
            except Exception as e:
                error_details.append(f"Row {idx + 5}: {str(e)}")
                error_count += 1
        
        return {
            "imported": imported_count,
            "errors": error_count,
            "error_details": error_details[:10],  # Limit to first 10 errors
            "message": f"Validated {imported_count} barangays. Note: To import actual measurement data, use the Admin interface to add individual child records."
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")

